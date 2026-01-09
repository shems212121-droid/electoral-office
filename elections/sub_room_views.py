from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from .models import SubOperationRoom, Introducer, Anchor, CenterDirector, PoliticalEntityAgent, Voter
from .sub_room_forms import SubOperationRoomForm, SubRoomFilterForm, AssignToRoomForm
import json


# ==================== Room Management Views ====================

@login_required
def sub_room_list(request):
    """عرض قائمة غرف العمليات الفرعية"""
    rooms = SubOperationRoom.objects.all().order_by('room_code')
    
    # Add statistics to each room
    rooms_with_stats = []
    for room in rooms:
        rooms_with_stats.append({
            'room': room,
            'introducers_count': room.get_introducers_count(),
            'voters_count': room.get_voters_count(),
            'directors_count': room.get_directors_count(),
            'agents_count': room.get_agents_count(),
            'total_count': room.get_total_people_count()
        })
    
    context = {
        'rooms_with_stats': rooms_with_stats,
        'total_rooms': rooms.count(),
        'active_rooms': rooms.filter(is_active=True).count()
    }
    
    return render(request, 'elections/sub_rooms/sub_room_list.html', context)


@login_required
def sub_room_create(request):
    """إضافة غرفة عمليات جديدة"""
    if request.method == 'POST':
        form = SubOperationRoomForm(request.POST)
        if form.is_valid():
            room = form.save()
            messages.success(request, f'تم إنشاء الغرفة {room.room_code} - {room.name} بنجاح!')
            return redirect('sub_room_dashboard', pk=room.pk)
    else:
        form = SubOperationRoomForm()
    
    context = {
        'form': form,
        'title': 'إضافة غرفة عمليات جديدة'
    }
    
    return render(request, 'elections/sub_rooms/sub_room_form.html', context)


@login_required
def sub_room_update(request, pk):
    """تعديل غرفة عمليات"""
    room = get_object_or_404(SubOperationRoom, pk=pk)
    
    if request.method == 'POST':
        form = SubOperationRoomForm(request.POST, instance=room)
        if form.is_valid():
            room = form.save()
            messages.success(request, f'تم تحديث الغرفة {room.room_code} بنجاح!')
            return redirect('sub_room_dashboard', pk=room.pk)
    else:
        form = SubOperationRoomForm(instance=room)
    
    context = {
        'form': form,
        'room': room,
        'title': f'تعديل الغرفة {room.room_code}'
    }
    
    return render(request, 'elections/sub_rooms/sub_room_form.html', context)


@login_required
def sub_room_delete(request, pk):
    """حذف غرفة عمليات"""
    room = get_object_or_404(SubOperationRoom, pk=pk)
    
    if request.method == 'POST':
        room_code = room.room_code
        room_name = room.name
        room.delete()
        messages.success(request, f'تم حذف الغرفة {room_code} - {room_name} بنجاح!')
        return redirect('sub_room_list')
    
    context = {
        'room': room,
        'introducers_count': room.get_introducers_count(),
        'voters_count': room.get_voters_count(),
        'directors_count': room.get_directors_count(),
        'agents_count': room.get_agents_count()
    }
    
    return render(request, 'elections/sub_rooms/sub_room_confirm_delete.html', context)


@login_required
def sub_room_toggle_status(request, pk):
    """تفعيل/تعطيل غرفة عمليات"""
    room = get_object_or_404(SubOperationRoom, pk=pk)
    room.is_active = not room.is_active
    room.save()
    
    status_text = 'تفعيل' if room.is_active else 'تعطيل'
    messages.success(request, f'تم {status_text} الغرفة {room.room_code} بنجاح!')
    
    return redirect('sub_room_list')


# ==================== Room Dashboard Views ====================

@login_required
def sub_room_dashboard(request, pk):
    """لوحة تحكم الغرفة"""
    room = get_object_or_404(SubOperationRoom, pk=pk)
    
    # Get all related data
    introducers = room.introducers.all()[:10]  # Latest 10
    anchors = room.anchors.all()[:10]
    directors = room.directors.all()[:10]
    agents = room.agents.all()[:10]
    
    # Statistics
    stats = {
        'introducers_count': room.get_introducers_count(),
        'voters_count': room.get_voters_count(),
        'anchors_count': room.get_anchors_count(),
        'directors_count': room.get_directors_count(),
        'agents_count': room.get_agents_count(),
        'total_count': room.get_total_people_count()
    }
    
    context = {
        'room': room,
        'stats': stats,
        'introducers': introducers,
        'anchors': anchors,
        'directors': directors,
        'agents': agents
    }
    
    return render(request, 'elections/sub_rooms/sub_room_dashboard.html', context)


@login_required
def sub_room_introducers(request, pk):
    """عرض المعرفين التابعين للغرفة"""
    room = get_object_or_404(SubOperationRoom, pk=pk)
    introducers = room.introducers.all().select_related('anchor', 'anchor__candidate')
    
    # Pagination
    paginator = Paginator(introducers, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'room': room,
        'introducers': page_obj,
        'total_count': introducers.count()
    }
    
    return render(request, 'elections/sub_rooms/sub_room_introducers.html', context)


@login_required
def sub_room_voters(request, pk):
    """عرض الناخبين التابعين للغرفة"""
    room = get_object_or_404(SubOperationRoom, pk=pk)
    voters = Voter.objects.filter(introducer__sub_room=room).select_related('introducer')
    
    # Pagination
    paginator = Paginator(voters, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'room': room,
        'voters': page_obj,
        'total_count': voters.count()
    }
    
    return render(request, 'elections/sub_rooms/sub_room_voters.html', context)


@login_required
def sub_room_directors(request, pk):
    """عرض مدراء المراكز التابعين للغرفة"""
    room = get_object_or_404(SubOperationRoom, pk=pk)
    directors = room.directors.all()
    
    # Pagination
    paginator = Paginator(directors, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'room': room,
        'directors': page_obj,
        'total_count': directors.count()
    }
    
    return render(request, 'elections/sub_rooms/sub_room_directors.html', context)


@login_required
def sub_room_agents(request, pk):
    """عرض وكلاء الكيان التابعين للغرفة"""
    room = get_object_or_404(SubOperationRoom, pk=pk)
    agents = room.agents.all().select_related('political_entity')
    
    # Pagination
    paginator = Paginator(agents, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'room': room,
        'agents': page_obj,
        'total_count': agents.count()
    }
    
    return render(request, 'elections/sub_rooms/sub_room_agents.html', context)


# ==================== Statistics and Reports ====================

@login_required
def sub_room_statistics(request):
    """إحصائيات تفصيلية لجميع الغرف"""
    rooms = SubOperationRoom.objects.all().order_by('room_code')
    
    statistics = []
    for room in rooms:
        statistics.append({
            'room': room,
            'introducers': room.get_introducers_count(),
            'voters': room.get_voters_count(),
            'anchors': room.get_anchors_count(),
            'directors': room.get_directors_count(),
            'agents': room.get_agents_count(),
            'total': room.get_total_people_count()
        })
    
    # Overall totals
    totals = {
        'introducers': sum(s['introducers'] for s in statistics),
        'voters': sum(s['voters'] for s in statistics),
        'anchors': sum(s['anchors'] for s in statistics),
        'directors': sum(s['directors'] for s in statistics),
        'agents': sum(s['agents'] for s in statistics),
        'total': sum(s['total'] for s in statistics)
    }
    
    context = {
        'statistics': statistics,
        'totals': totals,
        'rooms_count': rooms.count()
    }
    
    return render(request, 'elections/sub_rooms/sub_room_statistics.html', context)


@login_required
def sub_room_comparison(request):
    """مقارنة بين جميع الغرف"""
    rooms = SubOperationRoom.objects.all().order_by('room_code')
    
    comparison_data = []
    for room in rooms:
        comparison_data.append({
            'room_code': room.room_code,
            'room_name': room.name,
            'is_active': room.is_active,
            'supervisor': room.supervisor.username if room.supervisor else '-',
            'introducers': room.get_introducers_count(),
            'voters': room.get_voters_count(),
            'anchors': room.get_anchors_count(),
            'directors': room.get_directors_count(),
            'agents': room.get_agents_count(),
            'total': room.get_total_people_count()
        })
    
    context = {
        'comparison_data': comparison_data,
        'rooms_count': len(comparison_data)
    }
    
    return render(request, 'elections/sub_rooms/sub_room_comparison.html', context)


@login_required
def sub_room_export(request, pk):
    """تصدير بيانات الغرفة إلى Excel"""
    room = get_object_or_404(SubOperationRoom, pk=pk)
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Room info sheet
        ws_info = wb.create_sheet('معلومات الغرفة')
        ws_info.append(['كود الغرفة', room.room_code])
        ws_info.append(['اسم الغرفة', room.name])
        ws_info.append(['المشرف', room.supervisor.username if room.supervisor else '-'])
        ws_info.append(['الحالة', 'نشط' if room.is_active else 'غير نشط'])
        ws_info.append([])
        ws_info.append(['الإحصائيات'])
        ws_info.append(['عدد المعرفين', room.get_introducers_count()])
        ws_info.append(['عدد الناخبين', room.get_voters_count()])
        ws_info.append(['عدد مدراء المراكز', room.get_directors_count()])
        ws_info.append(['عدد وكلاء الكيان', room.get_agents_count()])
        ws_info.append(['المجموع الكلي', room.get_total_people_count()])
        
        # Introducers sheet
        ws_introducers = wb.create_sheet('المعرفين')
        ws_introducers.append(['كود المعرف', 'الاسم', 'رقم الهاتف', 'المرتكز', 'المرشح'])
        for introducer in room.introducers.all():
            ws_introducers.append([
                introducer.introducer_code,
                introducer.full_name,
                introducer.phone or '-',
                introducer.anchor.full_name if introducer.anchor else '-',
                introducer.anchor.candidate.full_name if introducer.anchor and introducer.anchor.candidate else '-'
            ])
        
        # Directors sheet
        ws_directors = wb.create_sheet('مدراء المراكز')
        ws_directors.append(['كود المدير', 'الاسم', 'رقم الهاتف', 'رقم المركز', 'اسم المركز'])
        for director in room.directors.all():
            ws_directors.append([
                director.director_code or '-',
                director.full_name,
                director.phone,
                director.assigned_center_number,
                director.assigned_center_name
            ])
        
        # Agents sheet
        ws_agents = wb.create_sheet('وكلاء الكيان')
        ws_agents.append(['كود الوكيل', 'الاسم', 'رقم الهاتف', 'الكيان السياسي', 'رقم المركز'])
        for agent in room.agents.all():
            ws_agents.append([
                agent.agent_code or '-',
                agent.full_name,
                agent.phone,
                agent.political_entity.name,
                agent.assigned_center_number
            ])
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{room.room_code}_data.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة. الرجاء تثبيتها أولاً.')
        return redirect('sub_room_dashboard', pk=pk)
