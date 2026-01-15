from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView, ListView, CreateView, DetailView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView, LogoutView
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Sum, Case, When, Value, IntegerField
from django.contrib import messages
from datetime import datetime, timedelta
from django.utils import timezone
from django.db import transaction
from django.contrib.auth.decorators import login_required

from .models import (
    Voter, Candidate, Anchor, Introducer, CandidateMonitor,
    CommunicationLog, CampaignTask, Area, Neighborhood,
    PoliticalParty, PartyCandidate, PollingCenter, PollingStation, VoteCount,
    UserRole, UserProfile, CivilSocietyObserver, InternationalObserver, PoliticalEntityAgent,
    SubOperationRoom
)
from .models_legacy import PersonHD, PCHd, VrcHD, GovernorateHD
from .forms import (
    CandidateForm, AnchorForm, IntroducerForm, VoterAssignmentForm,
    CommunicationLogForm, CampaignTaskForm, CandidateMonitorForm,
    PoliticalPartyForm, PartyCandidateForm, PollingCenterForm, 
    PollingStationForm, VoteCountForm, QuickVoteCountForm,
    GeneralVoteCountForm, SpecialVoteCountForm
)
from .decorators import role_required, permission_required, admin_only, can_export, can_delete


# ==================== PWA Offline Page ====================

def offline_page(request):
    """Offline page for PWA"""
    return render(request, 'elections/offline.html')


# ==================== Authentication ====================

class CustomLoginView(LoginView):
    template_name = 'elections/login.html'
    success_url = reverse_lazy('dashboard')
    redirect_authenticated_user = True


# ==================== Dashboard ====================

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'elections/dashboard.html'
    
    # Cache voter count (updated periodically) - Basra only
    CACHED_VOTER_COUNT = 1868927  # Pre-calculated from database (Basra governorate only)

    def get_context_data(self, **kwargs):
        from django.core.cache import cache
        from django.conf import settings
        
        context = super().get_context_data(**kwargs)
        
        try:
           # Try to get cached statistics
            cache_key = 'dashboard_stats'
            cached_stats = cache.get(cache_key)
            
            if cached_stats:
                # Use cached data
                context.update(cached_stats)
            else:
                # Calculate statistics (expensive queries)
                stats = {}
                
                # Basic Statistics
                try:
                    stats['total_candidates'] = Candidate.objects.count()
                    stats['total_anchors'] = Anchor.objects.count()
                    stats['total_introducers'] = Introducer.objects.count()
                    stats['total_sub_rooms'] = SubOperationRoom.objects.count()
                except Exception as e:
                    stats['db_error'] = str(e)
                    stats['total_candidates'] = 0
                    stats['total_anchors'] = 0
                    stats['total_introducers'] = 0
                    stats['total_sub_rooms'] = 0

                stats['total_voters'] = self.CACHED_VOTER_COUNT
                
                # Use raw SQL for faster counting with single query
                from django.db import connection
                try:
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            SELECT 
                                SUM(CASE WHEN introducer_id IS NOT NULL THEN 1 ELSE 0 END) as assigned,
                                SUM(CASE WHEN classification = 'supporter' THEN 1 ELSE 0 END) as supporters,
                                SUM(CASE WHEN classification = 'neutral' THEN 1 ELSE 0 END) as neutrals,
                                SUM(CASE WHEN classification = 'opponent' THEN 1 ELSE 0 END) as opponents
                            FROM elections_voter
                            WHERE introducer_id IS NOT NULL OR classification != 'unknown'
                            LIMIT 100000
                        """)
                        row = cursor.fetchone()
                        if row:
                            stats['assigned_voters'] = row[0] or 0
                            stats['supporter_count'] = row[1] or 0
                            stats['neutral_count'] = row[2] or 0
                            stats['opponent_count'] = row[3] or 0
                        else:
                            stats['assigned_voters'] = 0
                            stats['supporter_count'] = 0
                            stats['neutral_count'] = 0
                            stats['opponent_count'] = 0
                except Exception as db_e:
                     # Fallback for empty DB or migration issues
                    stats['assigned_voters'] = 0
                    stats['supporter_count'] = 0
                    stats['neutral_count'] = 0
                    stats['opponent_count'] = 0
                    stats['db_error_raw'] = str(db_e)
                
                stats['unknown_count'] = self.CACHED_VOTER_COUNT - (stats['supporter_count'] + stats['neutral_count'] + stats['opponent_count'])
                
                # Communication Statistics
                try:
                    stats['total_communications'] = CommunicationLog.objects.count()
                    stats['today_communications'] = CommunicationLog.objects.filter(
                        created_at__date=timezone.now().date()
                    ).count()
                    
                    # Task Statistics
                    stats['total_tasks'] = CampaignTask.objects.count()
                    stats['pending_tasks'] = CampaignTask.objects.filter(status='pending').count()
                    stats['in_progress_tasks'] = CampaignTask.objects.filter(status='in_progress').count()
                    stats['completed_tasks'] = CampaignTask.objects.filter(status='completed').count()
                except:
                    stats['total_communications'] = 0
                    stats['today_communications'] = 0
                    stats['total_tasks'] = 0
                    
                
                if stats.get('total_tasks', 0) > 0:
                    stats['task_completion_rate'] = (stats['completed_tasks'] / stats['total_tasks'] * 100)
                else:
                    stats['task_completion_rate'] = 0
                
                # Coverage Percentage
                if stats['total_voters'] > 0:
                    stats['coverage_percentage'] = (stats['assigned_voters'] / stats['total_voters'] * 100)
                else:
                    stats['coverage_percentage'] = 0
                
                # Cache for 5 minutes
                cache_timeout = getattr(settings, 'DASHBOARD_CACHE_TIMEOUT', 300)
                cache.set(cache_key, stats, cache_timeout)
                
                context.update(stats)
            
            # Recent Activities (not cached - always fresh)
            try:
                context['recent_communications'] = CommunicationLog.objects.select_related('caller').order_by('-created_at')[:10]
                context['recent_tasks'] = CampaignTask.objects.order_by('-created_at')[:5]
            except:
                context['recent_communications'] = []
                context['recent_tasks'] = []
            
            # Top Areas - Skip this slow query, use empty list
            context['top_areas'] = []
            
            return context
        except Exception as e:
            # Emergency context if everything fails
            return {
                'error_message': f'حدث خطأ غير متوقع في لوحة التحكم: {str(e)}',
                'total_voters': 0,
                'total_candidates': 0,
                'total_anchors': 0,
                'total_introducers': 0,
                'total_sub_rooms': 0,
                'recent_communications': [],
                'recent_tasks': [],
                'top_areas': [],
                'supporter_count': 0,
                'neutral_count': 0,
                'opponent_count': 0,
                'unknown_count': 0,
            }
        
    # def get_context_data(self, **kwargs):
    #     from django.core.cache import cache
    #     from django.conf import settings
        
    #     context = super().get_context_data(**kwargs)
        
    #     try:
    #        # Try to get cached statistics
    #         cache_key = 'dashboard_stats'
    #         cached_stats = cache.get(cache_key)
            
    #         if cached_stats:
    #             # Use cached data
    #             context.update(cached_stats)
    #         else:
    #             # Calculate statistics (expensive queries)
    #             stats = {}
                
    #             # Basic Statistics
    #             try:
    #                 stats['total_candidates'] = Candidate.objects.count()
    #                 stats['total_anchors'] = Anchor.objects.count()
    #                 stats['total_introducers'] = Introducer.objects.count()
    #                 stats['total_sub_rooms'] = SubOperationRoom.objects.count()
    #             except Exception as e:
    #                 stats['db_error'] = str(e)
    #                 stats['total_candidates'] = 0
    #                 stats['total_anchors'] = 0
    #                 stats['total_introducers'] = 0
    #                 stats['total_sub_rooms'] = 0

    #             stats['total_voters'] = self.CACHED_VOTER_COUNT
                
    #             # Use raw SQL for faster counting with single query
    #             from django.db import connection
    #             try:
    #                 with connection.cursor() as cursor:
    #                     cursor.execute("""
    #                         SELECT 
    #                             SUM(CASE WHEN introducer_id IS NOT NULL THEN 1 ELSE 0 END) as assigned,
    #                             SUM(CASE WHEN classification = 'supporter' THEN 1 ELSE 0 END) as supporters,
    #                             SUM(CASE WHEN classification = 'neutral' THEN 1 ELSE 0 END) as neutrals,
    #                             SUM(CASE WHEN classification = 'opponent' THEN 1 ELSE 0 END) as opponents
    #                         FROM elections_voter
    #                         WHERE introducer_id IS NOT NULL OR classification != 'unknown'
    #                         LIMIT 100000
    #                     """)
    #                     row = cursor.fetchone()
    #                     if row:
    #                         stats['assigned_voters'] = row[0] or 0
    #                         stats['supporter_count'] = row[1] or 0
    #                         stats['neutral_count'] = row[2] or 0
    #                         stats['opponent_count'] = row[3] or 0
    #                     else:
    #                         stats['assigned_voters'] = 0
    #                         stats['supporter_count'] = 0
    #                         stats['neutral_count'] = 0
    #                         stats['opponent_count'] = 0
    #             except Exception as db_e:
    #                  # Fallback for empty DB or migration issues
    #                 stats['assigned_voters'] = 0
    #                 stats['supporter_count'] = 0
    #                 stats['neutral_count'] = 0
    #                 stats['opponent_count'] = 0
    #                 stats['db_error_raw'] = str(db_e)
                
    #             stats['unknown_count'] = self.CACHED_VOTER_COUNT - (stats['supporter_count'] + stats['neutral_count'] + stats['opponent_count'])
                
    #             # Communication Statistics
    #             try:
    #                 stats['total_communications'] = CommunicationLog.objects.count()
    #                 stats['today_communications'] = CommunicationLog.objects.filter(
    #                     created_at__date=timezone.now().date()
    #                 ).count()
                    
    #                 # Task Statistics
    #                 stats['total_tasks'] = CampaignTask.objects.count()
    #                 stats['pending_tasks'] = CampaignTask.objects.filter(status='pending').count()
    #                 stats['in_progress_tasks'] = CampaignTask.objects.filter(status='in_progress').count()
    #                 stats['completed_tasks'] = CampaignTask.objects.filter(status='completed').count()
    #             except:
    #                 stats['total_communications'] = 0
    #                 stats['today_communications'] = 0
    #                 stats['total_tasks'] = 0
                    
                
    #             if stats.get('total_tasks', 0) > 0:
    #                 stats['task_completion_rate'] = (stats['completed_tasks'] / stats['total_tasks'] * 100)
    #             else:
    #                 stats['task_completion_rate'] = 0
                
    #             # Coverage Percentage
    #             if stats['total_voters'] > 0:
    #                 stats['coverage_percentage'] = (stats['assigned_voters'] / stats['total_voters'] * 100)
    #             else:
    #                 stats['coverage_percentage'] = 0
                
    #             # Cache for 5 minutes
    #             cache_timeout = getattr(settings, 'DASHBOARD_CACHE_TIMEOUT', 300)
    #             cache.set(cache_key, stats, cache_timeout)
                
    #             context.update(stats)
            
    #         # Recent Activities (not cached - always fresh)
    #         try:
    #             context['recent_communications'] = CommunicationLog.objects.select_related('caller').order_by('-created_at')[:10]
    #             context['recent_tasks'] = CampaignTask.objects.order_by('-created_at')[:5]
    #         except:
    #             context['recent_communications'] = []
    #             context['recent_tasks'] = []
            
    #         # Top Areas - Skip this slow query, use empty list
    #         context['top_areas'] = []
            
    #         return context
    #     except Exception as e:
    #         # Emergency context if everything fails
    #         return {
    #             'error_message': f'حدث خطأ غير متوقع في لوحة التحكم: {str(e)}',
    #             'total_voters': 0,
    #             'total_candidates': 0,
    #             'total_anchors': 0,
    #             'total_introducers': 0,
    #             'total_sub_rooms': 0,
    #             'recent_communications': [],
    #             'recent_tasks': [],
    #             'top_areas': [],
    #             'supporter_count': 0,
    #             'neutral_count': 0,
    #             'opponent_count': 0,
    #             'unknown_count': 0,
    #         }


# ==================== Voter Views ====================

class VoterListView(LoginRequiredMixin, ListView):
    model = Voter
    template_name = 'elections/voter_list.html'
    context_object_name = 'voters'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get('q')
        classification = self.request.GET.get('classification')
        has_introducer = self.request.GET.get('has_introducer')
        
        if q:
            queryset = queryset.filter(
                Q(voter_number__icontains=q) |
                Q(full_name__icontains=q) |
                Q(phone__icontains=q)
            )
        
        if classification:
            queryset = queryset.filter(classification=classification)
        
        if has_introducer == 'yes':
            queryset = queryset.filter(introducer__isnull=False)
        elif has_introducer == 'no':
            queryset = queryset.filter(introducer__isnull=True)
        
        return queryset.select_related('introducer', 'area')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_count'] = self.get_queryset().count()
        return context


class VoterDetailView(LoginRequiredMixin, DetailView):
    model = Voter
    template_name = 'elections/voter_detail.html'
    context_object_name = 'voter'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['communication_logs'] = self.object.logs.select_related('user').order_by('-created_at')
        context['log_form'] = CommunicationLogForm()
        return context


def voter_lookup_ajax(request):
    """AJAX endpoint to fetch voter data by voter_number from Legacy DB or Local DB"""
    from django.conf import settings
    
    voter_number = request.GET.get('voter_number')
    
    # Normalize Arabic numerals
    if voter_number:
        replacements = {
            '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
            '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9',
            '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
            '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9'
        }
        for arabic, english in replacements.items():
            voter_number = voter_number.replace(arabic, english)
        voter_number = voter_number.strip()

    introducer_id = request.GET.get('introducer_id')  # New parameter
    
    if not voter_number:
        return JsonResponse({'error': 'رقم الناخب مطلوب'}, status=400)
    
    # Check if voter exists in local database and is assigned to an introducer
    already_assigned = False
    same_introducer = False
    local_voter = None
    
    try:
        local_voter = Voter.objects.get(voter_number=voter_number)
        if local_voter.introducer:
            if introducer_id and str(local_voter.introducer.pk) == str(introducer_id):
                same_introducer = True
            else:
                already_assigned = True
    except Voter.DoesNotExist:
        local_voter = None
    
    # Check if legacy database is configured
    legacy_db_available = 'legacy_voters_db' in settings.DATABASES
    
    if legacy_db_available:
        try:
            # Query Legacy DB
            person = PersonHD.objects.using('legacy_voters_db').get(per_id=voter_number)
            
            # Get Center Name
            try:
                center = PCHd.objects.using('legacy_voters_db').get(pcno=person.pcno)
                center_name = center.pc_name
            except PCHd.DoesNotExist:
                center_name = ''

            # Get Registration Center Name
            try:
                vrc = VrcHD.objects.using('legacy_voters_db').get(vrc_id=person.per_vrc_id)
                reg_center_name = vrc.vrc_name_ar
            except (VrcHD.DoesNotExist, AttributeError):
                reg_center_name = ''

            # Get Governorate Name
            try:
                gov = GovernorateHD.objects.using('legacy_voters_db').get(gov_no=str(person.per_gov_id))
                gov_name = gov.gov_name
            except (GovernorateHD.DoesNotExist, AttributeError, ValueError):
                gov_name = ''

            # Construct Full Name
            full_name = filter(None, [person.per_first, person.per_father, person.per_grand])
            full_name = ' '.join(full_name)
            
            data = {
                'full_name': full_name,
                'date_of_birth': person.per_dob or '',
                'mother_name': '',
                'phone': '',
                'voting_center_number': person.pcno,
                'voting_center_name': center_name,
                'family_number': person.per_famno,
                'registration_center_name': reg_center_name,
                'registration_center_number': person.per_vrc_id,
                'governorate': gov_name, 
                'station_number': person.psno,
                'status': 'active',
                'found': True,
                'already_assigned': already_assigned,
                'same_introducer': same_introducer,
            }
            return JsonResponse(data)
        except PersonHD.DoesNotExist:
            pass  # Fall through to local voter check
        except Exception as e:
            # Database connection error or other issue - fall through to local check
            import logging
            logging.warning(f"Legacy DB lookup failed: {e}")
    
    # Fallback to local Voter model
    if local_voter:
        data = {
            'full_name': local_voter.full_name,
            'date_of_birth': local_voter.date_of_birth.strftime('%Y-%m-%d') if local_voter.date_of_birth else '',
            'mother_name': local_voter.mother_name or '',
            'phone': local_voter.phone or '',
            'voting_center_number': local_voter.voting_center_number or '',
            'voting_center_name': local_voter.voting_center_name or '',
            'family_number': local_voter.family_number or '',
            'registration_center_name': local_voter.registration_center_name or '',
            'registration_center_number': local_voter.registration_center_number or '',
            'governorate': local_voter.governorate or '',
            'station_number': local_voter.station_number or '',
            'status': local_voter.status or 'active',
            'found': True,
            'already_assigned': already_assigned,
            'same_introducer': same_introducer,
        }
        return JsonResponse(data)
    
    # Not found in any database
    total_voters_count = Voter.objects.count()
    
    # Debug file existence
    import os
    zip_exists = os.path.exists('voter_batches.zip')
    dir_exists = os.path.exists('voter_batches')
    batch_files_count = len(os.listdir('voter_batches')) if dir_exists else 0
    
    return JsonResponse({
        'found': False, 
        'error': 'الناخب غير موجود في قاعدة البيانات',
        'message': 'يرجى التأكد من صحة رقم الناخب',
        'debug_total_voters': total_voters_count,
        'debug_zip_exists': zip_exists,
        'debug_dir_exists': dir_exists,
        'debug_batch_files_count': batch_files_count
    })


# ==================== Candidate Views ====================

class CandidateListView(LoginRequiredMixin, ListView):
    model = Candidate
    template_name = 'elections/candidate_list.html'
    context_object_name = 'candidates'
    paginate_by = 20


class CandidateDetailView(LoginRequiredMixin, DetailView):
    model = Candidate
    template_name = 'elections/candidate_detail.html'
    context_object_name = 'candidate'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['anchors'] = self.object.anchors.all()
        context['monitors'] = self.object.monitors.all()
        context['anchors_count'] = self.object.get_anchors_count()
        context['introducers_count'] = self.object.get_introducers_count()
        context['voters_count'] = self.object.get_voters_count()
        context['monitors_count'] = self.object.get_monitors_count()
        return context


class CandidateCreateView(LoginRequiredMixin, CreateView):
    model = Candidate
    form_class = CandidateForm
    template_name = 'elections/candidate_form.html'
    success_url = reverse_lazy('candidate_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'تم إضافة المرشح بنجاح! الكود: {form.instance.candidate_code}')
        return super().form_valid(form)


# ==================== Anchor Views ====================

class AnchorListView(LoginRequiredMixin, ListView):
    model = Anchor
    template_name = 'elections/anchor_list.html'
    context_object_name = 'anchors'
    paginate_by = 30
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('candidate')
        candidate_id = self.request.GET.get('candidate')
        if candidate_id:
            queryset = queryset.filter(candidate_id=candidate_id)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['candidates'] = Candidate.objects.all()
        return context


class AnchorCreateView(LoginRequiredMixin, CreateView):
    model = Anchor
    form_class = AnchorForm
    template_name = 'elections/anchor_form.html'
    success_url = reverse_lazy('anchor_list')
    
    def form_valid(self, form):
        try:
            response = super().form_valid(form)
            messages.success(self.request, f'تم إضافة المرتكز بنجاح! الكود: {form.instance.anchor_code}')
            return response
        except Exception as e:
            from django.db import IntegrityError
            if isinstance(e, IntegrityError):
                if 'phone' in str(e):
                    form.add_error('phone', 'رقم الهاتف مسجل مسبقاً. يرجى استخدام رقم هاتف آخر.')
                elif 'voter_number' in str(e):
                    form.add_error('voter_number', 'رقم الناخب مسجل كمرتكز مسبقاً.')
                else:
                    messages.error(self.request, f'خطأ في قاعدة البيانات: {str(e)}')
            else:
                messages.error(self.request, f'حدث خطأ أثناء الحفظ: {str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        # عرض رسالة خطأ عامة مع تفاصيل الأخطاء
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                field_label = form.fields.get(field).label if field in form.fields else field
                for error in errors:
                    error_messages.append(f'{field_label}: {error}')
        
        if error_messages:
            messages.error(self.request, 'يوجد أخطاء في النموذج: ' + ' | '.join(error_messages))
        
        return super().form_invalid(form)


class AnchorDetailView(LoginRequiredMixin, DetailView):
    model = Anchor
    template_name = 'elections/anchor_detail.html'
    context_object_name = 'anchor'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['introducers'] = self.object.introducers.all()
        context['introducers_count'] = self.object.get_introducers_count()
        context['voters_count'] = self.object.get_voters_count()
        return context


class AnchorUpdateView(LoginRequiredMixin, UpdateView):
    model = Anchor
    form_class = AnchorForm
    template_name = 'elections/anchor_form.html'
    success_url = reverse_lazy('anchor_list')
    
    def form_valid(self, form):
        try:
            response = super().form_valid(form)
            messages.success(self.request, f'تم تحديث بيانات المرتكز: {form.instance.full_name}')
            return response
        except Exception as e:
            from django.db import IntegrityError
            if isinstance(e, IntegrityError):
                if 'phone' in str(e):
                    form.add_error('phone', 'رقم الهاتف مسجل مسبقاً. يرجى استخدام رقم هاتف آخر.')
                elif 'voter_number' in str(e):
                    form.add_error('voter_number', 'رقم الناخب مسجل كمرتكز مسبقاً.')
                else:
                    messages.error(self.request, f'خطأ في قاعدة البيانات: {str(e)}')
            else:
                messages.error(self.request, f'حدث خطأ أثناء الحفظ: {str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                field_label = form.fields.get(field).label if field in form.fields else field
                for error in errors:
                    error_messages.append(f'{field_label}: {error}')
        
        if error_messages:
            messages.error(self.request, 'يوجد أخطاء في النموذج: ' + ' | '.join(error_messages))
        
        return super().form_invalid(form)


class AnchorDeleteView(LoginRequiredMixin, DeleteView):
    model = Anchor
    template_name = 'elections/anchor_confirm_delete.html'
    success_url = reverse_lazy('anchor_list')
    context_object_name = 'anchor'
    
    def delete(self, request, *args, **kwargs):
        anchor = self.get_object()
        anchor_name = anchor.full_name
        try:
            # التحقق من وجود معرفين أو ناخبين مرتبطين
            introducers_count = anchor.introducers.count()
            voters_count = Voter.objects.filter(introducer__anchor=anchor).count()
            
            if introducers_count > 0:
                messages.error(request, f'لا يمكن حذف المرتكز {anchor_name} لأنه يحتوي على {introducers_count} معرف(ين)')
                return redirect('anchor_detail', pk=anchor.pk)
            
            if voters_count > 0:
                messages.error(request, f'لا يمكن حذف المرتكز {anchor_name} لأنه مرتبط بـ {voters_count} ناخب(ين)')
                return redirect('anchor_detail', pk=anchor.pk)
            
            messages.success(request, f'تم حذف المرتكز: {anchor_name}')
            return super().delete(request, *args, **kwargs)
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء الحذف: {str(e)}')
            return redirect('anchor_detail', pk=anchor.pk)


# ==================== Introducer Views ====================

class IntroducerListView(LoginRequiredMixin, ListView):
    model = Introducer
    template_name = 'elections/introducer_list.html'
    context_object_name = 'introducers'
    paginate_by = 30
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('anchor', 'anchor__candidate')
        anchor_id = self.request.GET.get('anchor')
        if anchor_id:
            queryset = queryset.filter(anchor_id=anchor_id)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['anchors'] = Anchor.objects.all()
        return context


class IntroducerCreateView(LoginRequiredMixin, CreateView):
    model = Introducer
    form_class = IntroducerForm
    template_name = 'elections/introducer_form.html'
    success_url = reverse_lazy('introducer_list')
    
    def form_valid(self, form):
        try:
            response = super().form_valid(form)
            messages.success(self.request, f'تم إضافة المعرف بنجاح! الكود: {form.instance.introducer_code}')
            return response
        except Exception as e:
            from django.db import IntegrityError
            if isinstance(e, IntegrityError):
                if 'phone' in str(e):
                    form.add_error('phone', 'رقم الهاتف مسجل مسبقاً. يرجى استخدام رقم هاتف آخر.')
                elif 'voter_number' in str(e):
                    form.add_error('voter_number', 'رقم الناخب مسجل كمعرف مسبقاً.')
                else:
                    messages.error(self.request, f'خطأ في قاعدة البيانات: {str(e)}')
            else:
                messages.error(self.request, f'حدث خطأ أثناء الحفظ: {str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                field_label = form.fields.get(field).label if field in form.fields else field
                for error in errors:
                    error_messages.append(f'{field_label}: {error}')
        
        if error_messages:
            messages.error(self.request, 'يوجد أخطاء في النموذج: ' + ' | '.join(error_messages))
        
        return super().form_invalid(form)


class IntroducerDetailView(LoginRequiredMixin, DetailView):
    model = Introducer
    template_name = 'elections/introducer_detail.html'
    context_object_name = 'introducer'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['voters'] = self.object.voters.all()[:50]  # Limited to 50 for performance
        context['voters_count'] = self.object.get_voters_count()
        return context


class IntroducerUpdateView(LoginRequiredMixin, UpdateView):
    model = Introducer
    form_class = IntroducerForm
    template_name = 'elections/introducer_form.html'
    success_url = reverse_lazy('introducer_list')
    
    def form_valid(self, form):
        try:
            response = super().form_valid(form)
            messages.success(self.request, f'تم تحديث بيانات المعرف: {form.instance.full_name}')
            return response
        except Exception as e:
            from django.db import IntegrityError
            if isinstance(e, IntegrityError):
                if 'phone' in str(e):
                    form.add_error('phone', 'رقم الهاتف مسجل مسبقاً. يرجى استخدام رقم هاتف آخر.')
                elif 'voter_number' in str(e):
                    form.add_error('voter_number', 'رقم الناخب مسجل كمعرف مسبقاً.')
                else:
                    messages.error(self.request, f'خطأ في قاعدة البيانات: {str(e)}')
            else:
                messages.error(self.request, f'حدث خطأ أثناء الحفظ: {str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                field_label = form.fields.get(field).label if field in form.fields else field
                for error in errors:
                    error_messages.append(f'{field_label}: {error}')
        
        if error_messages:
            messages.error(self.request, 'يوجد أخطاء في النموذج: ' + ' | '.join(error_messages))
        
        return super().form_invalid(form)


class IntroducerDeleteView(LoginRequiredMixin, DeleteView):
    model = Introducer
    template_name = 'elections/introducer_confirm_delete.html'
    success_url = reverse_lazy('introducer_list')
    context_object_name = 'introducer'
    
    def delete(self, request, *args, **kwargs):
        introducer = self.get_object()
        introducer_name = introducer.full_name
        try:
            # التحقق من وجود ناخبين مرتبطين
            voters_count = introducer.voters.count()
            
            if voters_count > 0:
                messages.error(request, f'لا يمكن حذف المعرف {introducer_name} لأنه مرتبط بـ {voters_count} ناخب(ين)')
                return redirect('introducer_detail', pk=introducer.pk)
            
            messages.success(request, f'تم حذف المعرف: {introducer_name}')
            return super().delete(request, *args, **kwargs)
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء الحذف: {str(e)}')
            return redirect('introducer_detail', pk=introducer.pk)


# ==================== Task Views ====================

class TaskListView(LoginRequiredMixin, ListView):
    model = CampaignTask
    template_name = 'elections/task_list.html'
    context_object_name = 'tasks'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_tasks = CampaignTask.objects.all()
        
        context['pending_tasks'] = all_tasks.filter(status='pending').order_by('-created_at')
        context['in_progress_tasks'] = all_tasks.filter(status='in_progress').order_by('-created_at')
        context['completed_tasks'] = all_tasks.filter(status='completed').order_by('-created_at')
        context['cancelled_tasks'] = all_tasks.filter(status='cancelled').order_by('-created_at')
        
        return context


class TaskCreateView(LoginRequiredMixin, CreateView):
    model = CampaignTask
    form_class = CampaignTaskForm
    template_name = 'elections/task_form.html'
    success_url = reverse_lazy('task_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'تم إنشاء المهمة بنجاح')
        return super().form_valid(form)


class TaskUpdateView(LoginRequiredMixin, UpdateView):
    model = CampaignTask
    form_class = CampaignTaskForm
    template_name = 'elections/task_form.html'
    success_url = reverse_lazy('task_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث المهمة بنجاح')
        return super().form_valid(form)


class TaskDeleteView(LoginRequiredMixin, DeleteView):
    model = CampaignTask
    template_name = 'elections/task_confirm_delete.html'
    success_url = reverse_lazy('task_list')
    context_object_name = 'task'
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'تم حذف المهمة بنجاح')
        return super().delete(request, *args, **kwargs)


@login_required
def task_change_status(request, pk):
    """تغيير حالة المهمة"""
    task = get_object_or_404(CampaignTask, pk=pk)
    new_status = request.POST.get('status')
    
    if new_status in ['pending', 'in_progress', 'completed', 'cancelled']:
        task.status = new_status
        task.save()
        
        status_names = {
            'pending': 'معلقة',
            'in_progress': 'قيد التنفيذ',
            'completed': 'مكتملة',
            'cancelled': 'ملغاة'
        }
        
        messages.success(request, f'تم تغيير حالة المهمة إلى: {status_names.get(new_status)}')
    else:
        messages.error(request, 'حالة غير صحيحة')
    
    return redirect('task_list')


# ==================== Communication ====================

def log_communication(request, voter_pk):
    """Add communication log for a voter"""
    voter = get_object_or_404(Voter, pk=voter_pk)
    
    if request.method == 'POST':
        form = CommunicationLogForm(request.POST)
        if form.is_valid():
            log = form.save(commit=False)
            log.voter = voter
            log.user = request.user
            log.save()
            messages.success(request, 'تم تسجيل الاتصال بنجاح')
        else:
            messages.error(request, 'خطأ في تسجيل الاتصال')
    
    return redirect('voter_detail', pk=voter_pk)


# ==================== Monitor Views ====================

class MonitorListView(LoginRequiredMixin, ListView):
    model = CandidateMonitor
    template_name = 'elections/monitor_list.html'
    context_object_name = 'monitors'
    paginate_by = 30
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('candidate')
        candidate_id = self.request.GET.get('candidate')
        status = self.request.GET.get('status')
        
        if candidate_id:
            queryset = queryset.filter(candidate_id=candidate_id)
        if status:
            queryset = queryset.filter(status=status)
            
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['candidates'] = Candidate.objects.all()
        context['total_count'] = self.get_queryset().count()
        return context


class MonitorCreateView(LoginRequiredMixin, CreateView):
    model = CandidateMonitor
    form_class = CandidateMonitorForm
    template_name = 'elections/monitor_form.html'
    success_url = reverse_lazy('monitor_list')
    
    def form_valid(self, form):
        form.instance.added_by = self.request.user
        messages.success(self.request, 'تم إضافة المراقب بنجاح')
        return super().form_valid(form)


class MonitorDetailView(LoginRequiredMixin, DetailView):
    model = CandidateMonitor
    template_name = 'elections/monitor_detail.html'
    context_object_name = 'monitor'


class MonitorUpdateView(LoginRequiredMixin, UpdateView):
    model = CandidateMonitor
    form_class = CandidateMonitorForm
    template_name = 'elections/monitor_form.html'
    success_url = reverse_lazy('monitor_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'تم تحديث بيانات المراقب بنجاح!')
        return super().form_valid(form)


class MonitorDeleteView(LoginRequiredMixin, DeleteView):
    model = CandidateMonitor
    template_name = 'elections/monitor_confirm_delete.html'
    success_url = reverse_lazy('monitor_list')
    context_object_name = 'monitor'
    
    def delete(self, request, *args, **kwargs):
        monitor = self.get_object()
        messages.success(request, f'تم حذف المراقب بنجاح!')
        return super().delete(request, *args, **kwargs)


# ==================== Search ====================

class VoterSearchView(LoginRequiredMixin, TemplateView):
    template_name = 'elections/voter_search.html'


# ==================== Vote Counting Views ====================

from elections.models import PoliticalParty, PartyCandidate, PollingCenter, PollingStation, VoteCount
from elections.forms import (
    PoliticalPartyForm, PartyCandidateForm, PollingCenterForm, 
    PollingStationForm, VoteCountForm, QuickVoteCountForm
)
from django.db.models import Sum, Count, Q


# Party Views
class PartyListView(LoginRequiredMixin, ListView):
    model = PoliticalParty
    template_name = 'elections/vote/party_list.html'
    context_object_name = 'parties'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_parties'] = PoliticalParty.objects.count()
        return context


class PartyCreateView(LoginRequiredMixin, CreateView):
    model = PoliticalParty
    form_class = PoliticalPartyForm
    template_name = 'elections/vote/party_form.html'
    success_url = reverse_lazy('party_list')


# Candidate Views
class PartyCandidateListView(LoginRequiredMixin, ListView):
    model = PartyCandidate
    template_name = 'elections/vote/candidate_list.html'
    context_object_name = 'candidates'
    paginate_by = 50
    
    def get_queryset(self):
        qs = PartyCandidate.objects.select_related('party').annotate(
            is_sadiqoun=Case(
                When(party__name__contains="الصادقون", then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        ).order_by('is_sadiqoun', 'party__serial_number', 'serial_number')
        
        # Filter by party
        party_id = self.request.GET.get('party')
        if party_id:
            qs = qs.filter(party_id=party_id)
        
        # Search by name, voter number, or phone
        query = self.request.GET.get('q', '').strip()
        if query:
            qs = qs.filter(
                Q(full_name__icontains=query) |
                Q(voter_number__icontains=query) |
                Q(phone__icontains=query)
            )
        
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['parties'] = PoliticalParty.objects.all()
        context['search_query'] = self.request.GET.get('q', '')
        return context


class PartyCandidateCreateView(LoginRequiredMixin, CreateView):
    model = PartyCandidate
    form_class = PartyCandidateForm
    template_name = 'elections/vote/candidate_form.html'
    success_url = reverse_lazy('candidate_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'تم إضافة المرشح بنجاح: {form.instance.full_name}')
        return super().form_valid(form)


class PartyCandidateDetailView(LoginRequiredMixin, DetailView):
    model = PartyCandidate
    template_name = 'elections/vote/candidate_detail.html'
    context_object_name = 'candidate'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_votes'] = self.object.get_total_votes()
        return context


class PartyCandidateUpdateView(LoginRequiredMixin, UpdateView):
    model = PartyCandidate
    form_class = PartyCandidateForm
    template_name = 'elections/vote/candidate_form.html'
    success_url = reverse_lazy('candidate_list')
    
    def form_valid(self, form):
        # Handle photo deletion if checkbox is checked
        if self.request.POST.get('clear_photo'):
            if form.instance.photo:
                # Delete the old photo file
                form.instance.photo.delete(save=False)
                form.instance.photo = None
        
        messages.success(self.request, f'تم تحديث بيانات المرشح: {form.instance.full_name}')
        return super().form_valid(form)


class PartyCandidateDeleteView(LoginRequiredMixin, TemplateView):
    template_name = 'elections/vote/candidate_confirm_delete.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['candidate'] = get_object_or_404(PartyCandidate, pk=kwargs['pk'])
        return context
    
    def post(self, request, pk):
        candidate = get_object_or_404(PartyCandidate, pk=pk)
        candidate_name = candidate.full_name
        candidate.delete()
        messages.success(request, f'تم حذف المرشح: {candidate_name}')
        return redirect('candidate_list')


# Polling Center Views
class PollingCenterListView(LoginRequiredMixin, ListView):
    model = PollingCenter
    template_name = 'elections/vote/center_list.html'
    context_object_name = 'centers'
    paginate_by = 30


class PollingCenterCreateView(LoginRequiredMixin, CreateView):
    model = PollingCenter
    form_class = PollingCenterForm
    template_name = 'elections/vote/center_form.html'
    success_url = reverse_lazy('polling_center_list')


# Vote Counting - Quick Entry
class QuickVoteCountView(LoginRequiredMixin, TemplateView):
    template_name = 'elections/vote/quick_count.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['stations'] = PollingStation.objects.select_related('center').all()
        context['parties'] = PoliticalParty.objects.prefetch_related('candidates').all()
        context['form'] = QuickVoteCountForm()
        return context
    
    def post(self, request, *args, **kwargs):
        station_id = request.POST.get('station')
        if not station_id:
            messages.error(request, 'يرجى اختيار المحطة')
            return redirect('quick_vote_count')
        
        station = get_object_or_404(PollingStation, id=station_id)
        candidates = PartyCandidate.objects.all()
        
        with transaction.atomic():
            for candidate in candidates:
                field_name = f'candidate_{candidate.id}'
                vote_count = request.POST.get(field_name, 0)
                try:
                    vote_count = int(vote_count) if vote_count else 0
                except ValueError:
                    vote_count = 0
                
                if vote_count > 0:
                    VoteCount.objects.update_or_create(
                        station=station,
                        candidate=candidate,
                        defaults={
                            'vote_count': vote_count,
                            'entered_by': request.user
                        }
                    )
        
        messages.success(request, f'تم حفظ جرد الأصوات للمحطة {station.full_number}')
        return redirect('quick_vote_count')


# Results Dashboard
class ResultsDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'elections/vote/results_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Party results
        party_results = PoliticalParty.objects.annotate(
            total_votes=Sum('candidates__vote_counts__vote_count')
        ).order_by('-total_votes')
        
        # Candidate results
        candidate_results = PartyCandidate.objects.annotate(
            total_votes=Sum('vote_counts__vote_count')
        ).select_related('party').order_by('-total_votes')[:20]
        
        # Statistics
        context['party_results'] = party_results
        context['candidate_results'] = candidate_results
        context['total_votes'] = VoteCount.objects.aggregate(Sum('vote_count'))['vote_count__sum'] or 0
        context['stations_counted'] = PollingStation.objects.filter(counting_status='completed').count()
        context['total_stations'] = PollingStation.objects.count()
        
        return context


# ==================== تحسينات المرشحين (Candidate Enhancements) ====================

def party_candidate_search_ajax(request):
    """AJAX endpoint للبحث الفوري في المرشحين"""
    query = request.GET.get('q', '').strip()
    party_id = request.GET.get('party', '').strip()
    
    if not query and not party_id:
        return JsonResponse({'results': []})
    
    # البحث في المرشحين
    candidates = PartyCandidate.objects.select_related('party', 'voter').all()
    
    if query:
        candidates = candidates.filter(
            Q(full_name__icontains=query) |
            Q(voter_number__icontains=query) |
            Q(phone__icontains=query) |
            Q(party__name__icontains=query)
        )
    
    if party_id:
        candidates = candidates.filter(party_id=party_id)
    
    # الحد الأقصى 50 نتيجة
    candidates = candidates[:50]
    
    results = []
    for candidate in candidates:
        results.append({
            'id': candidate.id,
            'full_name': candidate.full_name,
            'party_name': candidate.party.name,
            'party_number': candidate.party.serial_number,
            'serial_number': candidate.serial_number,
            'voter_number': candidate.voter_number or '-',
            'phone': candidate.phone or '-',
            'total_votes': candidate.get_total_votes(),
        })
    
    return JsonResponse({'results': results, 'count': len(results)})


class CandidateDashboardView(LoginRequiredMixin, TemplateView):
    """لوحة معلومات تفصيلية للمرشحين"""
    template_name = 'elections/vote/candidate_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # إحصائيات عامة
        context['total_candidates'] = PartyCandidate.objects.count()
        context['total_parties'] = PoliticalParty.objects.count()
        context['total_votes'] = VoteCount.objects.aggregate(Sum('vote_count'))['vote_count__sum'] or 0
        
        # ترتيب المرشحين حسب الأصوات
        candidates_with_votes = PartyCandidate.objects.annotate(
            total_votes=Sum('vote_counts__vote_count')
        ).select_related('party').order_by('-total_votes')[:20]
        
        context['top_candidates'] = candidates_with_votes
        
        # ترتيب الأحزاب حسب الأصوات
        parties_with_votes = PoliticalParty.objects.annotate(
            total_votes=Sum('candidates__vote_counts__vote_count'),
            candidates_count=Count('candidates')
        ).order_by('-total_votes')
        
        context['party_results'] = parties_with_votes
        
        # إحصائيات حسب الحزب
        party_stats = []
        for party in parties_with_votes:
            party_stats.append({
                'name': party.name,
                'serial_number': party.serial_number,
                'total_votes': party.total_votes or 0,
                'candidates_count': party.candidates_count,
                'percentage': (party.total_votes / context['total_votes'] * 100) if context['total_votes'] > 0 else 0
            })
        
        context['party_stats'] = party_stats
        
        # البيانات للرسوم البيانية (JSON)
        context['chart_labels'] = [p['name'] for p in party_stats[:10]]
        context['chart_data'] = [p['total_votes'] for p in party_stats[:10]]
        context['chart_colors'] = ['#667eea', '#764ba2', '#f093fb', '#4facfe', '#43e97b',
                                   '#fa709a', '#fee140', '#30cfd0', '#a8edea', '#fed6e3']
        
        return context



# ==================== Candidate Management Enhancements ====================

# AJAX Search
def candidate_search_ajax(request):
    """AJAX endpoint for instant candidate search"""
    query = request.GET.get('q', '').strip()
    party_id = request.GET.get('party', '')
    
    if not query and not party_id:
        return JsonResponse({'candidates': []})
    
    # Build query
    candidates = PartyCandidate.objects.select_related('party')
    
    if query:
        candidates = candidates.filter(
            Q(full_name__icontains=query) |
            Q(voter_number__icontains=query) |
            Q(phone__icontains=query)
        )
    
    if party_id:
        candidates = candidates.filter(party_id=party_id)
    
    # Limit results
    candidates = candidates[:50]
    
    # Format results
    results = []
    for c in candidates:
        results.append({
            'id': c.id,
            'full_name': c.full_name,
            'party': c.party.name,
            'party_id': c.party.id,
            'serial':f'{c.party.serial_number}-{c.serial_number}',
            'voter_number': c.voter_number or '-',
            'phone': c.phone or '-',
            'total_votes': c.get_total_votes(),
        })
    
    return JsonResponse({'candidates': results, 'count': len(results)})


# Export to Excel
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

@login_required
def export_candidates_excel(request):
    """Export candidates list to Excel"""
    # Get filters
    party_id = request.GET.get('party', '')
    search_query = request.GET.get('q', '')
    
    # Build queryset
    candidates = PartyCandidate.objects.select_related('party').order_by('party__serial_number', 'serial_number')
    
    if party_id:
        candidates = candidates.filter(party_id=party_id)
    
    if search_query:
        candidates = candidates.filter(
            Q(full_name__icontains=search_query) |
            Q(voter_number__icontains=search_query)
        )
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "المرشحون"
    
    # Header style
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = ['#', 'الرقم التعريفي', 'الاسم الكامل', 'الحزب/القائمة', 'رقم الناخب', 'الهاتف', 'إجمالي الأصوات']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data
    for idx, candidate in enumerate(candidates, 2):
        ws.cell(row=idx, column=1, value=idx-1)
        ws.cell(row=idx, column=2, value=f"{candidate.party.serial_number}-{candidate.serial_number}")
        ws.cell(row=idx, column=3, value=candidate.full_name)
        ws.cell(row=idx, column=4, value=candidate.party.name)
        ws.cell(row=idx, column=5, value=candidate.voter_number or '-')
        ws.cell(row=idx, column=6, value=candidate.phone or '-')
        ws.cell(row=idx, column=7, value=candidate.get_total_votes())
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="candidates_list.xlsx"'
    
    return response


# Export to PDF
@login_required
def export_candidates_pdf(request):
    """Export candidates list to PDF with Arabic support"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    try:
        # Register Arabic font (use a simple approach if no font available)
        # For production, you'd install and register an Arabic font
        pass
    except:
        pass
    
    # Get filters
    party_id = request.GET.get('party', '')
    search_query = request.GET.get('q', '')
    
    # Build queryset
    # Build queryset
    candidates = PartyCandidate.objects.select_related('party').annotate(
        is_sadiqoun=Case(
            When(party__name__contains="الصادقون", then=Value(0)),
            default=Value(1),
            output_field=IntegerField(),
        )
    ).order_by('is_sadiqoun', 'party__serial_number', 'serial_number')
    
    if party_id:
        candidates = candidates.filter(party_id=party_id)
    
    if search_query:
        candidates = candidates.filter(
            Q(full_name__icontains=search_query) |
            Q(voter_number__icontains=search_query)
        )
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    # Container for elements
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0066CC'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Table data
    data = [['#', 'ID', 'Name', 'Party', 'Voter #', 'Phone', 'Votes']]
    
    for idx, candidate in enumerate(candidates, 1):
        data.append([
            str(idx),
            f"{candidate.party.serial_number}-{candidate.serial_number}",
            candidate.full_name[:30],  # Truncate long names
            candidate.party.name[:20],
            candidate.voter_number or '-',
            candidate.phone or '-',
            str(candidate.get_total_votes())
        ])
    
    # Create table
    table = Table(data, colWidths=[0.6*inch, 1*inch, 2.5*inch, 1.8*inch, 1.2*inch, 1.2*inch, 0.8*inch])
    
    # Table style
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF from buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="candidates_list.pdf"'
    
    return response


# Candidates Dashboard
@login_required
def candidates_dashboard(request):
    """Statistics dashboard for candidates"""
    # General statistics
    total_candidates = PartyCandidate.objects.count()
    total_parties = PoliticalParty.objects.count()
    candidates_with_photos = PartyCandidate.objects.exclude(photo='').count()
    candidates_with_voters = PartyCandidate.objects.filter(voter__isnull=False).count()
    
    # Party distribution
    party_distribution = PoliticalParty.objects.annotate(
        candidate_count=Count('candidates')
    ).values('name', 'serial_number', 'candidate_count').order_by('serial_number')
    
    # Top candidates by votes
    top_candidates = PartyCandidate.objects.annotate(
        total_votes=Sum('vote_counts__vote_count')
    ).select_related('party').order_by('-total_votes')[:10]
    
    # Vote statistics by party
    party_votes = PoliticalParty.objects.annotate(
        total_votes=Sum('candidates__vote_counts__vote_count')
    ).order_by('-total_votes')
    
    context = {
        'total_candidates': total_candidates,
        'total_parties': total_parties,
        'candidates_with_photos': candidates_with_photos,
        'candidates_with_voters': candidates_with_voters,
        'candidates_without_photos': total_candidates - candidates_with_photos,
        'party_distribution': list(party_distribution),
        'top_candidates': top_candidates,
        'party_votes': party_votes,
    }
    
    return render(request, 'elections/vote/candidates_dashboard.html', context)


# ==================== Introducer Voters Management ====================

class IntroducerVotersView(LoginRequiredMixin, DetailView):
    """صفحة إدارة ناخبي المعرف"""
    model = Introducer
    template_name = 'elections/introducer_voters.html'
    context_object_name = 'introducer'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['voters'] = self.object.voters.all().order_by('-created_at')
        context['voters_count'] = self.object.voters.count()
        return context


@login_required
def add_voter_to_introducer(request, pk):
    """إضافة ناخب للمعرف عبر AJAX"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'}, status=405)
    
    introducer = get_object_or_404(Introducer, pk=pk)
    voter_number = request.POST.get('voter_number', '')
    
    # Normalize Arabic numerals
    if voter_number:
        replacements = {
            '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
            '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9',
            '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
            '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9'
        }
        for arabic, english in replacements.items():
            voter_number = voter_number.replace(arabic, english)
    
    voter_number = voter_number.strip()
    phone = request.POST.get('phone', '').strip()
    
    if not voter_number:
        return JsonResponse({'success': False, 'error': 'رقم الناخب مطلوب'})
    
    # البحث عن الناخب في قاعدة البيانات
    try:
        voter = Voter.objects.get(voter_number=voter_number)
        
        # التحقق إذا كان مرتبطاً بنفس المعرف
        if voter.introducer == introducer:
            return JsonResponse({'success': False, 'error': 'الناخب مسجل بالفعل لهذا المعرف'})
        
        # ربط الناخب بالمعرف
        voter.introducer = introducer
        if phone:
            voter.phone = phone
        
        # توليد كود الناخب
        count = Voter.objects.filter(introducer=introducer).count() + 1
        voter.voter_code = f"{introducer.introducer_code}-VOT-{count:03d}"
        
        voter.save()
        
        return JsonResponse({
            'success': True,
            'total_voters': introducer.voters.count(),
            'voter': {
                'id': voter.pk,
                'voter_number': voter.voter_number,
                'voter_code': voter.voter_code,
                'full_name': voter.full_name,
                'phone': voter.phone or '',
                'voting_center_name': voter.voting_center_name[:30] if voter.voting_center_name else '',
            }
        })
        
    except Voter.DoesNotExist:
        # محاولة إنشاء ناخب جديد من قاعدة البيانات الخارجية
        try:
            from .models_legacy import PersonHD, PCHd, VrcHD, GovernorateHD
            
            person = PersonHD.objects.using('legacy_voters_db').get(per_id=voter_number)
            
            # الحصول على اسم المركز
            try:
                center = PCHd.objects.using('legacy_voters_db').get(pcno=person.pcno)
                center_name = center.pc_name
            except PCHd.DoesNotExist:
                center_name = ''
            
            # الحصول على اسم مركز التسجيل
            try:
                vrc = VrcHD.objects.using('legacy_voters_db').get(vrc_id=person.per_vrc_id)
                reg_center_name = vrc.vrc_name_ar
            except:
                reg_center_name = ''
            
            # الحصول على اسم المحافظة
            try:
                gov = GovernorateHD.objects.using('legacy_voters_db').get(gov_no=str(person.per_gov_id))
                gov_name = gov.gov_name
            except:
                gov_name = 'البصرة'
            
            # بناء الاسم الكامل
            full_name = ' '.join(filter(None, [person.per_first, person.per_father, person.per_grand]))
            
            # إنشاء ناخب جديد
            count = Voter.objects.filter(introducer=introducer).count() + 1
            voter_code = f"{introducer.introducer_code}-VOT-{count:03d}"
            
            voter = Voter.objects.create(
                voter_number=voter_number,
                full_name=full_name,
                date_of_birth=person.per_dob,
                phone=phone or '',
                voting_center_number=person.pcno or '',
                voting_center_name=center_name,
                registration_center_name=reg_center_name,
                registration_center_number=person.per_vrc_id or '',
                governorate=gov_name,
                station_number=person.psno or '',
                family_number=person.per_famno or '',
                introducer=introducer,
                voter_code=voter_code,
            )
            
            return JsonResponse({
                'success': True,
                'total_voters': introducer.voters.count(),
                'voter': {
                    'id': voter.pk,
                    'voter_number': voter.voter_number,
                    'voter_code': voter.voter_code,
                    'full_name': voter.full_name,
                    'phone': voter.phone or '',
                    'voting_center_name': voter.voting_center_name[:30] if voter.voting_center_name else '',
                }
            })
            
        except PersonHD.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'الناخب غير موجود في قاعدة البيانات'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'خطأ: {str(e)}'})


@login_required
def remove_voter_from_introducer(request, pk):
    """إزالة ناخب من المعرف"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'}, status=405)
    
    introducer = get_object_or_404(Introducer, pk=pk)
    voter_id = request.POST.get('voter_id')
    
    if not voter_id:
        return JsonResponse({'success': False, 'error': 'معرف الناخب مطلوب'})
    
    try:
        voter = Voter.objects.get(pk=voter_id, introducer=introducer)
        voter.introducer = None
        voter.voter_code = None
        voter.save()
        
        return JsonResponse({
            'success': True,
            'total_voters': introducer.voters.count()
        })
    except Voter.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'الناخب غير موجود'})


@login_required
def bulk_add_voters_to_introducer(request, pk):
    """إضافة ناخبين متعددين للمعرف"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة غير مسموحة'}, status=405)
    
    introducer = get_object_or_404(Introducer, pk=pk)
    voter_numbers_str = request.POST.get('voter_numbers', '')
    
    if not voter_numbers_str:
        return JsonResponse({'success': False, 'error': 'أرقام الناخبين مطلوبة'})

    # Normalize Arabic numerals in the bulk string first
    replacements = {
        '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
        '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9',
        '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
        '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9'
    }
    for arabic, english in replacements.items():
        voter_numbers_str = voter_numbers_str.replace(arabic, english)
    
    voter_numbers = [n.strip() for n in voter_numbers_str.split(',') if n.strip()]
    
    added = 0
    already_exists = 0
    not_found = 0
    
    for voter_number in voter_numbers:
        try:
            voter = Voter.objects.get(voter_number=voter_number)
            
            if voter.introducer == introducer:
                already_exists += 1
                continue
            
            # ربط الناخب بالمعرف
            voter.introducer = introducer
            count = Voter.objects.filter(introducer=introducer).count() + 1
            voter.voter_code = f"{introducer.introducer_code}-VOT-{count:03d}"
            voter.save()
            added += 1
            
        except Voter.DoesNotExist:
            # محاولة إنشاء من قاعدة البيانات الخارجية
            try:
                from .models_legacy import PersonHD, PCHd
                
                person = PersonHD.objects.using('legacy_voters_db').get(per_id=voter_number)
                
                try:
                    center = PCHd.objects.using('legacy_voters_db').get(pcno=person.pcno)
                    center_name = center.pc_name
                except:
                    center_name = ''
                
                full_name = ' '.join(filter(None, [person.per_first, person.per_father, person.per_grand]))
                
                count = Voter.objects.filter(introducer=introducer).count() + 1
                voter_code = f"{introducer.introducer_code}-VOT-{count:03d}"
                
                Voter.objects.create(
                    voter_number=voter_number,
                    full_name=full_name,
                    voting_center_number=person.pcno or '',
                    voting_center_name=center_name,
                    station_number=person.psno or '',
                    family_number=person.per_famno or '',
                    introducer=introducer,
                    voter_code=voter_code,
                    governorate='البصرة',
                )
                added += 1
                
            except:
                not_found += 1
    
    return JsonResponse({
        'success': True,
        'added': added,
        'already_exists': already_exists,
        'not_found': not_found,
        'total_voters': introducer.voters.count()
    })


@login_required
def voter_lookup_for_introducer(request):
    """بحث عن ناخب مع التحقق من ارتباطه بمعرف"""
    voter_number = request.GET.get('voter_number', '')
    
    # Normalize Arabic numerals
    if voter_number:
        replacements = {
            '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
            '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9',
            '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
            '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9'
        }
        for arabic, english in replacements.items():
            voter_number = voter_number.replace(arabic, english)
    
    voter_number = voter_number.strip()
    introducer_id = request.GET.get('introducer_id')
    
    if not voter_number:
        return JsonResponse({'found': False, 'error': 'رقم الناخب مطلوب'})
    
    # البحث أولاً في قاعدة البيانات المحلية
    try:
        voter = Voter.objects.get(voter_number=voter_number)
        
        data = {
            'found': True,
            'full_name': voter.full_name,
            'date_of_birth': voter.date_of_birth.strftime('%Y-%m-%d') if voter.date_of_birth else '',
            'voting_center_name': voter.voting_center_name,
            'already_assigned': voter.introducer is not None,
            'same_introducer': str(voter.introducer_id) == str(introducer_id) if introducer_id else False,
            'current_introducer': voter.introducer.full_name if voter.introducer else None,
        }
        return JsonResponse(data)
        
    except Voter.DoesNotExist:
        # البحث في قاعدة البيانات الخارجية
        try:
            from .models_legacy import PersonHD, PCHd
            
            person = PersonHD.objects.using('legacy_voters_db').get(per_id=voter_number)
            
            try:
                center = PCHd.objects.using('legacy_voters_db').get(pcno=person.pcno)
                center_name = center.pc_name
            except:
                center_name = ''
            
            full_name = ' '.join(filter(None, [person.per_first, person.per_father, person.per_grand]))
            
            return JsonResponse({
                'found': True,
                'full_name': full_name,
                'date_of_birth': person.per_dob or '',
                'voting_center_name': center_name,
                'already_assigned': False,
                'same_introducer': False,
                'is_new': True,  # سيتم إنشاؤه عند الإضافة
            })
            
        except PersonHD.DoesNotExist:
            return JsonResponse({'found': False, 'error': 'الناخب غير موجود'})




# ==================== Data Reset View ====================
from django.views import View
from django.contrib import messages
from django.db import transaction

class DataResetView(LoginRequiredMixin, View):
    template_name = 'elections/data_reset.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.error(request, "عذراً، هذا الإجراء متاح للمشرف الأعلى فقط.")
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        if not request.user.is_superuser:
            return redirect('dashboard')

        reset_targets = request.POST.getlist('reset_target')
        deleted_counts = {}

        try:
            with transaction.atomic():
                if 'candidates' in reset_targets:
                    count, _ = Candidate.objects.all().delete()
                    pc_count, _ = PartyCandidate.objects.all().delete()
                    deleted_counts['المرشحين'] = count + pc_count

                if 'anchors' in reset_targets:
                    count, _ = Anchor.objects.all().delete()
                    deleted_counts['المرتكزات'] = count

                if 'introducers' in reset_targets:
                    count, _ = Introducer.objects.all().delete()
                    deleted_counts['المعرفين'] = count

                if 'voters' in reset_targets:
                    count, _ = Voter.objects.all().delete()
                    deleted_counts['الناخبين'] = count

                if 'observers' in reset_targets:
                    count1, _ = CandidateMonitor.objects.all().delete()
                    count2, _ = CivilSocietyObserver.objects.all().delete()
                    count3, _ = InternationalObserver.objects.all().delete()
                    count4, _ = PoliticalEntityAgent.objects.all().delete()
                    deleted_counts['المراقبين والوكلاء'] = count1 + count2 + count3 + count4

                if 'results_general' in reset_targets:
                    count, _ = VoteCount.objects.filter(vote_type='general').delete()
                    deleted_counts['نتائج التصويت العام'] = count

                if 'results_special' in reset_targets:
                    count, _ = VoteCount.objects.filter(vote_type='special').delete()
                    deleted_counts['نتائج التصويت الخاص'] = count

            if deleted_counts:
                msg_parts = [f"{k}: {v}" for k, v in deleted_counts.items()]
                messages.success(request, f"تم حذف البيانات بنجاح: {', '.join(msg_parts)}")
            else:
                messages.warning(request, "لم يتم تحديد أي بيانات للحذف.")

        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء حذف البيانات: {str(e)}")

        return redirect('data_reset')

# ==================== Background Sync APIs ====================

@login_required
def api_candidates_list(request):
    """المرشحين للمزامنة"""
    candidates = PartyCandidate.objects.all().values('id', 'full_name', 'candidate_number', 'party_id')
    return JsonResponse(list(candidates), safe=False)

@login_required
def api_parties_list(request):
    """الكيانات للمزامنة"""
    parties = PoliticalParty.objects.all().values('id', 'name', 'code')
    return JsonResponse(list(parties), safe=False)

# ==================== Emergency Import Tool ====================

def run_import_script(request):
    """Trigger the voter import script"""
    import threading
    from django.core.management import call_command
    
    # Only allow superusers OR secret key
    secret_key = request.GET.get('secret')
    if secret_key != 'shems_voter_import_2024_secure' and not request.user.is_superuser:
        return HttpResponse('Unauthorized - Admin Access Only', status=403)
        
    def run_in_background():
        import io
        from contextlib import redirect_stdout, redirect_stderr
        
        try:
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                import datetime
                f.write(f"\n[{datetime.datetime.now()}] --- Starting voter import process ---\n")
            
            output = io.StringIO()
            with redirect_stdout(output), redirect_stderr(output):
                call_command('import_voters')
            
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                f.write(output.getvalue())
                import datetime
                f.write(f"[{datetime.datetime.now()}] --- Voter import process ended ---\n")
        except Exception as e:
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                import datetime
                f.write(f"[{datetime.datetime.now()}] CRITICAL Voter import failure: {e}\n")

    # Start the thread
    thread = threading.Thread(target=run_in_background)
    thread.daemon = True
    thread.start()

    return HttpResponse(f'''
        <h1>✅ Import Started Successfully!</h1>
        <p>The voter import process is now running in the background.</p>
        <p>You can close this page and continue using the site.</p>
        <p>Data will appear gradually as it is processed (approx 10-20 mins).</p>
        <p><a href="/tool/import-log/">View Import Log</a></p>
        <p><a href="/dashboard/">Return to Dashboard</a></p>
    ''')

def view_import_log(request):
    """View the import log file"""
    import os
    log_file = 'import_log.txt'
    
    if not os.path.exists(log_file):
        return HttpResponse('Log file not found yet.', content_type='text/plain')
        
    with open(log_file, 'r', encoding='utf-8') as f:
        content = f.read()
        
    return HttpResponse(content, content_type='text/plain; charset=utf-8')

def run_link_hierarchy(request):
    """Trigger the link_electoral_hierarchy management command via URL"""
    # Security: check for secret key or admin access
    secret = request.GET.get('secret')
    if secret != 'shems_voter_import_2024_secure' and not request.user.is_superuser:
        return HttpResponse('Unauthorized - Admin Access Only', status=403)
        
    import threading
    from django.core.management import call_command
    
    def run_in_background():
        import io
        from contextlib import redirect_stdout, redirect_stderr
        
        try:
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                import datetime
                f.write(f"\n[{datetime.datetime.now()}] --- Starting electoral hierarchy linking process ---\n")
            
            output = io.StringIO()
            with redirect_stdout(output), redirect_stderr(output):
                # Apply migrations first to be safe
                print("Running migrations...")
                call_command('migrate', interactive=False)
                # Link hierarchy
                print("Linking electoral hierarchy...")
                call_command('link_electoral_hierarchy')
            
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                import datetime
                f.write(output.getvalue())
                f.write(f"[{datetime.datetime.now()}] --- Electoral hierarchy linking process ended ---\n")
        except Exception as e:
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                import datetime
                f.write(f"[{datetime.datetime.now()}] CRITICAL Hierarchy linking failure: {e}\n")

    # Start the thread
    thread = threading.Thread(target=run_in_background)
    thread.daemon = True
    thread.start()

    return HttpResponse(f'''
        <h1>✅ Hierarchy Linking Started Successfully!</h1>
        <p>The process of linking voters, stations, and centers is now running in the background.</p>
        <p>You can check the status in the <a href="/tool/import-log/">Import Log</a>.</p>
        <p><a href="/dashboard/">Return to Dashboard</a></p>
    ''')

def run_import_centers(request):
    """Trigger the import of general and special polling centers via URL"""
    secret = request.GET.get('secret')
    if secret != 'shems_voter_import_2024_secure' and not request.user.is_superuser:
        return HttpResponse('Unauthorized - Admin Access Only', status=403)
        
    import threading
    from django.core.management import call_command
    
    def run_in_background():
        import io
        from contextlib import redirect_stdout, redirect_stderr
        
        try:
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                import datetime
                f.write(f"\n[{datetime.datetime.now()}] --- Starting centers import ---\n")
            
            output = io.StringIO()
            with redirect_stdout(output), redirect_stderr(output):
                # Import General Centers
                try:
                    call_command('import_general_polling_centers')
                except Exception as ex:
                    print(f"General import exception: {ex}")
                
                # Import Special Centers
                try:
                    call_command('import_special_polling_centers')
                except Exception as ex:
                    print(f"Special import exception: {ex}")
            
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                import datetime
                f.write(output.getvalue())
                f.write(f"[{datetime.datetime.now()}] --- Centers import process ended ---\n")
        except Exception as e:
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                import datetime
                f.write(f"[{datetime.datetime.now()}] CRITICAL Centers import failure: {e}\n")

    # Debug: log view call
    debug_msg = ""
    try:
        import os
        cwd = os.getcwd()
        with open('import_log.txt', 'a', encoding='utf-8') as f:
            import datetime
            f.write(f"[{datetime.datetime.now()}] run_import_centers view triggered in {cwd}.\n")
        debug_msg = f"Successfully wrote to import_log.txt in {cwd}"
    except Exception as e:
        debug_msg = f"Failed to write log: {e} in {os.getcwd()}"

    # Start the thread
    thread = threading.Thread(target=run_in_background)
    thread.daemon = True
    thread.start()

    return HttpResponse(f'''
        <h1>✅ Centers Import Started Successfully!</h1>
        <p>Debug: {debug_msg}</p>
        <p>General and Special polling centers are being imported in the background.</p>
        <p>Check the <a href="/tool/import-log/">Import Log</a> for progress.</p>
        <p><a href="/dashboard/">Return to Dashboard</a></p>
    ''')

def run_import_part2(request):
    """Import Part 2: Batches 29-30"""
    return _run_import_subset(request, "Part 2 (29-30)", 29, 31)

def run_import_part3(request):
    """Import Part 3: Batches 31-38"""
    return _run_import_subset(request, "Part 3 (31-38)", 31, 39)

def _run_import_subset(request, name, start_batch, end_batch):
    """Helper to run a subset of batches"""
    secret = request.GET.get('secret')
    if secret != 'shems_voter_import_2024_secure' and not request.user.is_superuser:
        return HttpResponse('Unauthorized - Admin Access Only', status=403)
        
    import threading
    from django.core.management import call_command
    
    def run_in_background():
        import io
        from contextlib import redirect_stdout, redirect_stderr
        
        try:
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                import datetime
                f.write(f"\n[{datetime.datetime.now()}] --- Starting {name} import ---\n")
            
            output = io.StringIO()
            with redirect_stdout(output), redirect_stderr(output):
                # Call command with start/end arguments (needs update in management command)
                # But here we will rely on the script arguments update or environment variables
                # For simplicity, we'll set an ENV var that the script can read
                import os
                os.environ['IMPORT_START_BATCH'] = str(start_batch)
                os.environ['IMPORT_END_BATCH'] = str(end_batch)
                
                call_command('import_voters')
            
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                import datetime
                f.write(output.getvalue())
                f.write(f"[{datetime.datetime.now()}] --- {name} import process ended ---\n")
        except Exception as e:
            with open('import_log.txt', 'a', encoding='utf-8') as f:
                import datetime
                f.write(f"[{datetime.datetime.now()}] CRITICAL {name} import failure: {e}\n")

    # Start the thread
    thread = threading.Thread(target=run_in_background)
    thread.daemon = True
    thread.start()

    return HttpResponse(f'''
        <h1>✅ {name} Import Started!</h1>
        <p>Importing batches {start_batch} to {end_batch-1}.</p>
        <p>Check the <a href="/tool/import-log/">Import Log</a> for progress.</p>
        <p><a href="/dashboard/">Return to Dashboard</a></p>
    ''')
