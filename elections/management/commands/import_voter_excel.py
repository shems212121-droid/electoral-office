from django.core.management.base import BaseCommand, CommandError
from elections.models import Voter
from openpyxl import load_workbook
import os
from django.db import transaction


class Command(BaseCommand):
    help = 'استيراد بيانات الناخبين من ملفات Excel'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='مسار ملف Excel')

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        
        # التحقق من وجود الملف
        if not os.path.exists(excel_file):
            raise CommandError(f'الملف غير موجود: {excel_file}')

        self.stdout.write(self.style.SUCCESS(f'\n{"="*60}'))
        self.stdout.write(self.style.SUCCESS(f'بدء استيراد البيانات من: {excel_file}'))
        self.stdout.write(self.style.SUCCESS(f'{"="*60}\n'))

        # إحصائيات
        stats = {
            'total': 0,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0
        }

        try:
            # تحميل ملف Excel
            wb = load_workbook(excel_file, read_only=True)
            ws = wb.active
            
            # الحصول على أسماء الأعمدة
            headers = [cell.value for cell in ws[1]]
            self.stdout.write(f'الأعمدة المتوفرة: {headers}\n')
            
            # تحديد مواقع الأعمدة
            name_col = 0  # العمود الأول (اسم المعرف/المراقب)
            voter_num_col = 1  # رقم الناخب
            phone_col = 2  # رقم الهاتف
            center_num_col = 3  # رقم المركز
            
            # معالجة الصفوف
            rows = list(ws.rows)
            total_rows = len(rows) - 1  # استثناء صف الرأس
            
            # معالجة البيانات على دفعات
            batch_size = 500
            for batch_start in range(1, len(rows), batch_size):
                batch_end = min(batch_start + batch_size, len(rows))
                batch_rows = rows[batch_start:batch_end]
                
                with transaction.atomic():
                    for idx_in_batch, row in enumerate(batch_rows):
                        idx = batch_start + idx_in_batch
                        try:
                            # استخراج البيانات
                            name = row[name_col].value
                            voter_number = row[voter_num_col].value
                            phone = row[phone_col].value
                            center_number = row[center_num_col].value
                        
                            # التحقق من البيانات الأساسية
                            if not voter_number:
                                stats['skipped'] += 1
                                continue
                            
                            # تحويل البيانات للنص
                            voter_number = str(voter_number).strip()
                            name = str(name).strip() if name else ''
                            phone = str(phone).strip() if phone else ''
                            center_number = str(center_number).strip() if center_number else ''
                            
                            stats['total'] += 1
                            
                            # البحث عن اسم مركز الاقتراع من البيانات الموجودة
                            center_name = ''
                            if center_number:
                                existing_voter = Voter.objects.filter(
                                    voting_center_number=center_number
                                ).exclude(voting_center_name='').first()
                                
                                if existing_voter:
                                    center_name = existing_voter.voting_center_name
                            
                            # التحقق من وجود الناخب
                            voter, created = Voter.objects.update_or_create(
                                voter_number=voter_number,
                                defaults={
                                    'full_name': name,
                                    'phone': phone,
                                    'voting_center_number': center_number,
                                    'voting_center_name': center_name,
                                }
                            )
                            
                            if created:
                                stats['created'] += 1
                            else:
                                stats['updated'] += 1
                            
                            # عرض التقدم كل 1000 سجل
                            if idx % 1000 == 0:
                                progress = (idx / total_rows) * 100
                                self.stdout.write(
                                    f'التقدم: {idx}/{total_rows} ({progress:.1f}%) - '
                                    f'إضافة: {stats["created"]}, تحديث: {stats["updated"]}, '
                                    f'تخطي: {stats["skipped"]}, أخطاء: {stats["errors"]}'
                                )
                        
                        except Exception as e:
                            stats['errors'] += 1
                            self.stdout.write(
                                self.style.ERROR(f'خطأ في الصف {idx}: {str(e)}')
                            )
                            continue
            
            wb.close()
            
            # عرض التقرير النهائي
            self.stdout.write(self.style.SUCCESS(f'\n{"="*60}'))
            self.stdout.write(self.style.SUCCESS('تقرير الاستيراد النهائي'))
            self.stdout.write(self.style.SUCCESS(f'{"="*60}'))
            self.stdout.write(f'إجمالي الصفوف المعالجة: {stats["total"]}')
            self.stdout.write(self.style.SUCCESS(f'✓ تم الإضافة: {stats["created"]}'))
            self.stdout.write(self.style.WARNING(f'⟳ تم التحديث: {stats["updated"]}'))
            self.stdout.write(self.style.WARNING(f'⊘ تم التخطي: {stats["skipped"]}'))
            self.stdout.write(self.style.ERROR(f'✗ الأخطاء: {stats["errors"]}'))
            self.stdout.write(self.style.SUCCESS(f'{"="*60}\n'))
            
            # إحصائيات قاعدة البيانات
            total_voters = Voter.objects.count()
            self.stdout.write(
                self.style.SUCCESS(f'إجمالي الناخبين في قاعدة البيانات: {total_voters:,}')
            )
            
        except Exception as e:
            raise CommandError(f'خطأ في معالجة الملف: {str(e)}')
