"""
نظام التقارير الشاملة للمكتب الانتخابي
يوفر تقارير PDF و Excel و CSV بحجم A4
"""

from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Q
from datetime import datetime
from django.utils import timezone
import csv
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from .models import (
    Voter, Candidate, Anchor, Introducer, CommunicationLog, CampaignTask,
    PartyCandidate, PoliticalParty, PollingCenter, PollingStation, VoteCount,
    CenterDirector, CandidateMonitor, PoliticalEntityAgent
)
from archive.models import ArchivedDocument


# ==================== Helper Functions ====================

def get_excel_styles():
    """إرجاع أنماط Excel الموحدة"""
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    return {
        'header_fill': header_fill,
        'header_font': header_font,
        'header_alignment': header_alignment,
        'data_alignment': data_alignment,
        'border': thin_border
    }


def setup_excel_sheet(ws, title, headers, col_widths):
    """إعداد ورقة Excel بالتنسيق الموحد"""
    ws.title = title
    ws.right_to_left = True
    
    styles = get_excel_styles()
    
    # إضافة العناوين
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
    
    # ضبط عرض الأعمدة
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    return styles


# ==================== تقرير المرشحين الشامل ====================

@login_required
def export_candidates_comprehensive_excel(request):
    """تصدير تقرير المرشحين الشامل إلى Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    ws = wb.active
    
    headers = [
        'م', 'رقم الحزب', 'اسم الحزب', 'رقم المرشح', 'اسم المرشح',
        'اسم الأم', 'المواليد', 'رقم الهاتف',
        'الأصوات العامة', 'الأصوات الخاصة', 'إجمالي الأصوات',
        'محطات صوتت', 'محطات لم تصوت'
    ]
    col_widths = [5, 12, 35, 12, 40, 30, 15, 15, 15, 15, 15, 15, 15]
    
    styles = setup_excel_sheet(ws, "تقرير المرشحين", headers, col_widths)
    
    candidates = PartyCandidate.objects.select_related('party').order_by(
        'party__serial_number', 'serial_number'
    )
    
    for row_num, candidate in enumerate(candidates, 2):
        data = [
            row_num - 1,
            candidate.party.serial_number if candidate.party else '',
            candidate.party.name if candidate.party else 'مستقل',
            candidate.serial_number,
            candidate.full_name,
            candidate.mother_name_triple,
            candidate.date_of_birth,
            candidate.phone,
            candidate.get_general_votes(),
            candidate.get_special_votes(),
            candidate.get_total_votes(),
            candidate.stations_voted,
            candidate.stations_not_voted
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['data_alignment']
            cell.border = styles['border']
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="candidates_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


@login_required
def export_candidates_comprehensive_csv(request):
    """تصدير تقرير المرشحين الشامل إلى CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="candidates_report_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'م', 'رقم الحزب', 'اسم الحزب', 'رقم المرشح', 'اسم المرشح',
        'اسم الأم', 'المواليد', 'رقم الهاتف',
        'الأصوات العامة', 'الأصوات الخاصة', 'إجمالي الأصوات',
        'محطات صوتت', 'محطات لم تصوت'
    ])
    
    candidates = PartyCandidate.objects.select_related('party').order_by(
        'party__serial_number', 'serial_number'
    )
    
    for idx, candidate in enumerate(candidates, 1):
        writer.writerow([
            idx,
            candidate.party.serial_number if candidate.party else '',
            candidate.party.name if candidate.party else 'مستقل',
            candidate.serial_number,
            candidate.full_name,
            candidate.mother_name_triple,
            candidate.date_of_birth,
            candidate.phone,
            candidate.get_general_votes(),
            candidate.get_special_votes(),
            candidate.get_total_votes(),
            candidate.stations_voted,
            candidate.stations_not_voted
        ])
    
    return response


@login_required
def export_candidates_comprehensive_pdf(request):
    """تصدير تقرير المرشحين الشامل إلى PDF"""
    if not PDF_AVAILABLE:
        return HttpResponse('مكتبة reportlab غير متوفرة', status=500)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="candidates_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), 
                           leftMargin=1*cm, rightMargin=1*cm,
                           topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                  fontSize=16, alignment=TA_CENTER)
    
    elements.append(Paragraph("تقرير المرشحين الشامل", title_style))
    elements.append(Paragraph(f"التاريخ: {datetime.now().strftime('%Y-%m-%d')}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    table_data = [['م', 'الحزب', 'المرشح', 'عام', 'خاص', 'الإجمالي']]
    
    candidates = PartyCandidate.objects.select_related('party').order_by(
        'party__serial_number', 'serial_number'
    )
    
    for idx, c in enumerate(candidates, 1):
        party_num = c.party.serial_number if c.party else '---'
        table_data.append([
            str(idx),
            f"{party_num}",
            c.full_name[:30],
            str(c.get_general_votes()),
            str(c.get_special_votes()),
            str(c.get_total_votes())
        ])
    
    table = Table(table_data, colWidths=[1*cm, 2*cm, 8*cm, 2.5*cm, 2.5*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return response


# ==================== تقرير الناخبين الشامل ====================

@login_required
def export_voters_comprehensive_excel(request):
    """تصدير تقرير الناخبين الشامل إلى Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    ws = wb.active
    
    headers = [
        'م', 'رقم الناخب', 'الاسم الكامل', 'تاريخ الميلاد', 'اسم الأم',
        'المحافظة', 'رقم مركز الاقتراع', 'اسم مركز الاقتراع',
        'رقم المحطة', 'التصنيف', 'رقم الهاتف'
    ]
    col_widths = [6, 15, 35, 12, 25, 12, 15, 40, 10, 12, 15]
    
    styles = setup_excel_sheet(ws, "تقرير الناخبين", headers, col_widths)
    
    # تحديد عدد الناخبين للتصدير (أول 50000 لتجنب الحجم الكبير)
    voters = Voter.objects.all()[:50000]
    
    for row_num, voter in enumerate(voters, 2):
        data = [
            row_num - 1,
            voter.voter_number,
            voter.full_name,
            str(voter.date_of_birth) if voter.date_of_birth else '',
            voter.mother_name or '',
            voter.governorate,
            voter.voting_center_number,
            voter.voting_center_name,
            voter.station_number or '',
            voter.get_classification_display(),
            voter.phone or ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['data_alignment']
            cell.border = styles['border']
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="voters_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


@login_required
def export_voters_comprehensive_csv(request):
    """تصدير تقرير الناخبين إلى CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="voters_report_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'م', 'رقم الناخب', 'الاسم الكامل', 'تاريخ الميلاد', 'اسم الأم',
        'المحافظة', 'رقم مركز الاقتراع', 'اسم مركز الاقتراع',
        'رقم المحطة', 'التصنيف', 'رقم الهاتف'
    ])
    
    voters = Voter.objects.all()
    
    for idx, voter in enumerate(voters, 1):
        writer.writerow([
            idx,
            voter.voter_number,
            voter.full_name,
            str(voter.date_of_birth) if voter.date_of_birth else '',
            voter.mother_name or '',
            voter.governorate,
            voter.voting_center_number,
            voter.voting_center_name,
            voter.station_number or '',
            voter.get_classification_display(),
            voter.phone or ''
        ])
    
    return response


# ==================== تقرير المعرفين ====================

@login_required
def export_introducers_excel(request):
    """تصدير تقرير المعرفين إلى Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    ws = wb.active
    
    headers = [
        'م', 'كود المعرف', 'الاسم الكامل', 'رقم الناخب', 'رقم الهاتف',
        'المرتكز', 'المرشح', 'عدد الناخبين', 'مركز الاقتراع'
    ]
    col_widths = [5, 25, 35, 15, 15, 35, 35, 12, 40]
    
    styles = setup_excel_sheet(ws, "تقرير المعرفين", headers, col_widths)
    
    introducers = Introducer.objects.select_related('anchor', 'anchor__candidate').all()
    
    for row_num, introducer in enumerate(introducers, 2):
        anchor_name = introducer.anchor.full_name if introducer.anchor else '---'
        candidate_name = introducer.anchor.candidate.full_name if (introducer.anchor and introducer.anchor.candidate) else '---'
        
        data = [
            row_num - 1,
            introducer.introducer_code or '',
            introducer.full_name,
            introducer.voter_number,
            introducer.phone or '',
            anchor_name,
            candidate_name,
            introducer.get_voters_count(),
            introducer.voting_center_name or ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['data_alignment']
            cell.border = styles['border']
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="introducers_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


@login_required
def export_introducers_csv(request):
    """تصدير تقرير المعرفين إلى CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="introducers_report_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'م', 'كود المعرف', 'الاسم الكامل', 'رقم الناخب', 'رقم الهاتف',
        'المرتكز', 'المرشح', 'عدد الناخبين', 'مركز الاقتراع'
    ])
    
    introducers = Introducer.objects.select_related('anchor', 'anchor__candidate').all()
    
    for idx, introducer in enumerate(introducers, 1):
        anchor_name = introducer.anchor.full_name if introducer.anchor else '---'
        candidate_name = introducer.anchor.candidate.full_name if (introducer.anchor and introducer.anchor.candidate) else '---'
        
        writer.writerow([
            idx,
            introducer.introducer_code or '',
            introducer.full_name,
            introducer.voter_number,
            introducer.phone or '',
            anchor_name,
            candidate_name,
            introducer.get_voters_count(),
            introducer.voting_center_name or ''
        ])
    
    return response


# ==================== تقرير المرتكزات ====================

@login_required
def export_anchors_excel(request):
    """تصدير تقرير المرتكزات إلى Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    ws = wb.active
    
    headers = [
        'م', 'كود المرتكز', 'الاسم الكامل', 'رقم الناخب', 'رقم الهاتف',
        'المرشح', 'عدد المعرفين', 'عدد الناخبين', 'مركز الاقتراع'
    ]
    col_widths = [5, 25, 35, 15, 15, 35, 12, 12, 40]
    
    styles = setup_excel_sheet(ws, "تقرير المرتكزات", headers, col_widths)
    
    anchors = Anchor.objects.select_related('candidate').all()
    
    for row_num, anchor in enumerate(anchors, 2):
        candidate_name = anchor.candidate.full_name if anchor.candidate else '---'
        
        data = [
            row_num - 1,
            anchor.anchor_code or '',
            anchor.full_name,
            anchor.voter_number,
            anchor.phone or '',
            candidate_name,
            anchor.get_introducers_count(),
            anchor.get_voters_count(),
            anchor.voting_center_name or ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['data_alignment']
            cell.border = styles['border']
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="anchors_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


@login_required
def export_anchors_csv(request):
    """تصدير تقرير المرتكزات إلى CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="anchors_report_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'م', 'كود المرتكز', 'الاسم الكامل', 'رقم الناخب', 'رقم الهاتف',
        'المرشح', 'عدد المعرفين', 'عدد الناخبين', 'مركز الاقتراع'
    ])
    
    anchors = Anchor.objects.select_related('candidate').all()
    
    for idx, anchor in enumerate(anchors, 1):
        candidate_name = anchor.candidate.full_name if anchor.candidate else '---'
        
        writer.writerow([
            idx,
            anchor.anchor_code or '',
            anchor.full_name,
            anchor.voter_number,
            anchor.phone or '',
            candidate_name,
            anchor.get_introducers_count(),
            anchor.get_voters_count(),
            anchor.voting_center_name or ''
        ])
    
    return response


# ==================== تقرير جرد الأصوات ====================

@login_required
def export_votes_excel(request):
    """تصدير تقرير جرد الأصوات إلى Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    ws = wb.active
    
    headers = [
        'م', 'رقم المركز', 'اسم المركز', 'رقم المحطة', 'نوع التصويت',
        'رقم المرشح', 'اسم المرشح', 'عدد الأصوات'
    ]
    col_widths = [6, 12, 40, 12, 15, 12, 40, 12]
    
    styles = setup_excel_sheet(ws, "جرد الأصوات", headers, col_widths)
    
    votes = VoteCount.objects.select_related(
        'station', 'station__center', 'candidate', 'candidate__party'
    ).order_by('station__center__center_number', 'station__station_number')
    
    for row_num, vote in enumerate(votes, 2):
        data = [
            row_num - 1,
            vote.station.center.center_number,
            vote.station.center.name,
            vote.station.station_number,
            vote.get_vote_type_display(),
            f"{vote.candidate.party.serial_number if vote.candidate.party else '---'}-{vote.candidate.serial_number}",
            vote.candidate.full_name,
            vote.vote_count
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['data_alignment']
            cell.border = styles['border']
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="votes_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


@login_required
def export_votes_csv(request):
    """تصدير تقرير جرد الأصوات إلى CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="votes_report_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'م', 'رقم المركز', 'اسم المركز', 'رقم المحطة', 'نوع التصويت',
        'رقم المرشح', 'اسم المرشح', 'عدد الأصوات'
    ])
    
    votes = VoteCount.objects.select_related(
        'station', 'station__center', 'candidate', 'candidate__party'
    ).order_by('station__center__center_number', 'station__station_number')
    
    for idx, vote in enumerate(votes, 1):
        party_num = vote.candidate.party.serial_number if vote.candidate.party else '---'
        writer.writerow([
            idx,
            vote.station.center.center_number,
            vote.station.center.name,
            vote.station.station_number,
            vote.get_vote_type_display(),
            f"{party_num}-{vote.candidate.serial_number}",
            vote.candidate.full_name,
            vote.vote_count
        ])
    
    return response


# ==================== تقرير ملخص النتائج ====================

@login_required
def export_results_summary_excel(request):
    """تصدير ملخص نتائج التصويت إلى Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    
    # صفحة ملخص المرشحين
    ws1 = wb.active
    ws1.title = "ملخص المرشحين"
    ws1.right_to_left = True
    
    headers1 = ['الترتيب', 'رقم المرشح', 'اسم المرشح', 'الحزب', 'أصوات عامة', 'أصوات خاصة', 'الإجمالي']
    styles = setup_excel_sheet(ws1, "ملخص المرشحين", headers1, [8, 12, 40, 30, 12, 12, 12])
    
    candidates = PartyCandidate.objects.select_related('party').all()
    candidates_with_votes = []
    for c in candidates:
        candidates_with_votes.append({
            'candidate': c,
            'general': c.get_general_votes(),
            'special': c.get_special_votes(),
            'total': c.get_total_votes()
        })
    
    candidates_with_votes.sort(key=lambda x: x['total'], reverse=True)
    
    for row_num, item in enumerate(candidates_with_votes, 2):
        c = item['candidate']
        party_num = c.party.serial_number if c.party else '---'
        party_name = c.party.name if c.party else 'مستقل'
        
        data = [
            row_num - 1,
            f"{party_num}-{c.serial_number}",
            c.full_name,
            party_name,
            item['general'],
            item['special'],
            item['total']
        ]
        for col_num, value in enumerate(data, 1):
            cell = ws1.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['data_alignment']
            cell.border = styles['border']
    
    # صفحة ملخص الأحزاب
    ws2 = wb.create_sheet("ملخص الأحزاب")
    ws2.right_to_left = True
    
    headers2 = ['الترتيب', 'رقم الحزب', 'اسم الحزب', 'عدد المرشحين', 'إجمالي الأصوات']
    setup_excel_sheet(ws2, "ملخص الأحزاب", headers2, [8, 12, 40, 15, 15])
    
    parties = PoliticalParty.objects.all()
    parties_with_votes = []
    for p in parties:
        parties_with_votes.append({
            'party': p,
            'candidates': p.get_candidates_count(),
            'total': p.get_total_votes()
        })
    
    parties_with_votes.sort(key=lambda x: x['total'], reverse=True)
    
    for row_num, item in enumerate(parties_with_votes, 2):
        p = item['party']
        data = [row_num - 1, p.serial_number, p.name, item['candidates'], item['total']]
        for col_num, value in enumerate(data, 1):
            cell = ws2.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['data_alignment']
            cell.border = styles['border']
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="results_summary_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


# ==================== لوحة التقارير الشاملة ====================

@login_required
def comprehensive_reports_dashboard(request):
    """لوحة التقارير الشاملة"""
    
    # إحصائيات عامة
    total_voters = Voter.objects.count()
    total_candidates = PartyCandidate.objects.count()
    total_anchors = Anchor.objects.count()
    total_introducers = Introducer.objects.count()
    total_centers = PollingCenter.objects.count()
    total_stations = PollingStation.objects.count()
    total_votes = VoteCount.objects.aggregate(total=Sum('vote_count'))['total'] or 0
    total_parties = PoliticalParty.objects.count()
    
    context = {
        'excel_available': EXCEL_AVAILABLE,
        'pdf_available': PDF_AVAILABLE,
        'total_voters': total_voters,
        'total_candidates': total_candidates,
        'total_anchors': total_anchors,
        'total_introducers': total_introducers,
        'total_centers': total_centers,
        'total_stations': total_stations,
        'total_votes': total_votes,
        'total_parties': total_parties,
        'report_date': datetime.now(),
    }
    
    return render(request, 'elections/comprehensive_reports.html', context)


# ==================== تقرير مدراء المراكز ====================

@login_required
def export_center_directors_excel(request):
    """تصدير تقرير مدراء المراكز إلى Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    ws = wb.active
    
    headers = [
        'م', 'الاسم الكامل', 'نوع الاقتراع', 'اسم المركز المخصص', 'رقم المركز',
        'رقم الهاتف', 'عدد الوكلاء', 'عدد المحطات المغطاة', 'الحالة'
    ]
    col_widths = [5, 30, 12, 35, 12, 15, 12, 15, 10]
    
    styles = setup_excel_sheet(ws, "مدراء المراكز", headers, col_widths)
    
    directors = CenterDirector.objects.all().order_by('voting_type', 'assigned_center_number')
    
    for row_num, director in enumerate(directors, 2):
        data = [
            row_num - 1,
            director.full_name,
            director.get_voting_type_display(),
            director.assigned_center_name,
            director.assigned_center_number,
            director.phone,
            director.get_agents_count(),
            director.get_stations_with_agents(),
            director.get_status_display()
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['data_alignment']
            cell.border = styles['border']
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="center_directors_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response

@login_required
def export_center_directors_csv(request):
    """تصدير تقرير مدراء المراكز إلى CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="center_directors_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'م', 'الاسم الكامل', 'نوع الاقتراع', 'اسم المركز المخصص', 'رقم المركز',
        'رقم الهاتف', 'عدد الوكلاء', 'عدد المحطات المغطاة', 'الحالة'
    ])
    
    directors = CenterDirector.objects.all().order_by('voting_type', 'assigned_center_number')
    
    for idx, director in enumerate(directors, 1):
        writer.writerow([
            idx,
            director.full_name,
            director.get_voting_type_display(),
            director.assigned_center_name,
            director.assigned_center_number,
            director.phone,
            director.get_agents_count(),
            director.get_stations_with_agents(),
            director.get_status_display()
        ])
    
    return response


# ==================== تقرير المراقبين ====================

@login_required
def export_monitors_excel(request):
    """تصدير تقرير المراقبين إلى Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    ws = wb.active
    
    headers = [
        'م', 'الاسم الكامل', 'الدور', 'المرشح', 'رقم المركز',
        'اسم المركز', 'رقم المحطة', 'رقم الهاتف', 'الحالة'
    ]
    col_widths = [5, 30, 15, 30, 12, 35, 10, 15, 12]
    
    styles = setup_excel_sheet(ws, "المراقبين", headers, col_widths)
    
    monitors = CandidateMonitor.objects.select_related('candidate').all()
    
    for row_num, monitor in enumerate(monitors, 2):
        data = [
            row_num - 1,
            monitor.full_name,
            monitor.get_role_type_display(),
            monitor.candidate.full_name if monitor.candidate else '',
            monitor.voting_center_number,
            monitor.voting_center_name,
            monitor.station_number,
            monitor.phone,
            monitor.get_status_display()
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['data_alignment']
            cell.border = styles['border']
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="monitors_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response

@login_required
def export_monitors_csv(request):
    """تصدير تقرير المراقبين إلى CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="monitors_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'م', 'الاسم الكامل', 'الدور', 'المرشح', 'رقم المركز',
        'اسم المركز', 'رقم المحطة', 'رقم الهاتف', 'الحالة'
    ])
    
    monitors = CandidateMonitor.objects.select_related('candidate').all()
    
    for idx, monitor in enumerate(monitors, 1):
        writer.writerow([
            idx,
            monitor.full_name,
            monitor.get_role_type_display(),
            monitor.candidate.full_name if monitor.candidate else '',
            monitor.voting_center_number,
            monitor.voting_center_name,
            monitor.station_number,
            monitor.phone,
            monitor.get_status_display()
        ])
    
    return response


# ==================== تقرير الوكلاء ====================

@login_required
def export_agents_excel(request):
    """تصدير تقرير الوكلاء إلى Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    ws = wb.active
    
    headers = [
        'م', 'الاسم الكامل', 'الكيان السياسي', 'مدير المركز', 'مركز الاقتراع المخصص',
        'رقم المحطة المخصصة', 'رقم الهاتف', 'الحالة'
    ]
    col_widths = [5, 30, 25, 25, 35, 15, 15, 12]
    
    styles = setup_excel_sheet(ws, "الوكلاء", headers, col_widths)
    
    agents = PoliticalEntityAgent.objects.select_related('political_entity', 'center_director').all()
    
    for row_num, agent in enumerate(agents, 2):
        data = [
            row_num - 1,
            agent.full_name,
            agent.political_entity.name,
            agent.center_director.full_name if agent.center_director else '',
            agent.assigned_center_name,
            agent.assigned_station_number,
            agent.phone,
            agent.get_status_display()
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['data_alignment']
            cell.border = styles['border']
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="agents_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response

@login_required
def export_agents_csv(request):
    """تصدير تقرير الوكلاء إلى CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="agents_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'م', 'الاسم الكامل', 'الكيان السياسي', 'مدير المركز', 'مركز الاقتراع المخصص',
        'رقم المحطة المخصصة', 'رقم الهاتف', 'الحالة'
    ])
    
    agents = PoliticalEntityAgent.objects.select_related('political_entity', 'center_director').all()
    
    for idx, agent in enumerate(agents, 1):
        writer.writerow([
            idx,
            agent.full_name,
            agent.political_entity.name,
            agent.center_director.full_name if agent.center_director else '',
            agent.assigned_center_name,
            agent.assigned_station_number,
            agent.phone,
            agent.get_status_display()
        ])
    
    return response


# ==================== تقرير الأرشيف ====================

@login_required
def export_archive_excel(request):
    """تصدير تقرير الأرشيف إلى Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    ws = wb.active
    
    headers = [
        'م', 'العنوان', 'نوع الوثيقة', 'المجلد', 'تاريخ الوثيقة',
        'الكلمات المفتاحية', 'أرشفت بواسطة', 'تاريخ الأرشفة'
    ]
    col_widths = [5, 40, 20, 25, 15, 30, 20, 20]
    
    styles = setup_excel_sheet(ws, "الأرشيف", headers, col_widths)
    
    documents = ArchivedDocument.objects.select_related('folder').all()
    
    for row_num, doc in enumerate(documents, 2):
        data = [
            row_num - 1,
            doc.title,
            doc.document_type,
            str(doc.folder) if doc.folder else '',
            str(doc.document_date) if doc.document_date else '',
            doc.tags,
            doc.archived_by,
            doc.archived_at.strftime("%Y-%m-%d %H:%M")
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['data_alignment']
            cell.border = styles['border']
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="archive_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response

@login_required
def export_archive_csv(request):
    """تصدير تقرير الأرشيف إلى CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="archive_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'م', 'العنوان', 'نوع الوثيقة', 'المجلد', 'تاريخ الوثيقة',
        'الكلمات المفتاحية', 'أرشفت بواسطة', 'تاريخ الأرشفة'
    ])
    
    documents = ArchivedDocument.objects.select_related('folder').all()
    
    for idx, doc in enumerate(documents, 1):
        writer.writerow([
            idx,
            doc.title,
            doc.document_type,
            str(doc.folder) if doc.folder else '',
            str(doc.document_date) if doc.document_date else '',
            doc.tags,
            doc.archived_by,
            doc.archived_at.strftime("%Y-%m-%d %H:%M")
        ])
    
    return response

