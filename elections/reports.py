"""
نظام التقارير المتقدم للمكتب الانتخابي
يوفر وظائف توليد تقارير PDF و Excel
"""

from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from datetime import datetime, timedelta
from django.utils import timezone
import csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from .models import (
    Voter, Candidate, Anchor, Introducer, CommunicationLog, CampaignTask,
    PartyCandidate, PoliticalParty, PollingCenter, VoteCount
)


# ==================== CSV Export ====================

@login_required
def export_voters_csv(request):
    """تصدير الناخبين إلى CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="voters.csv"'
    response.write('\ufeff')  # UTF-8 BOM
    
    writer = csv.writer(response)
    writer.writerow([
        'رقم الناخب', 'الاسم الكامل', 'تاريخ الميلاد', 'اسم الأم',
        'رقم الهاتف', 'المحافظة', 'مركز الاقتراع', 'رقم المركز',
        'رقم المحطة', 'التصنيف', 'المعرّف', 'الحالة'
    ])
    
    voters = Voter.objects.select_related('introducer').all()
    
    for voter in voters:
        writer.writerow([
            voter.voter_number,
            voter.full_name,
            voter.date_of_birth or '',
            voter.mother_name or '',
            voter.phone or '',
            voter.governorate,
            voter.voting_center_name,
            voter.voting_center_number,
            voter.station_number or '',
            voter.get_classification_display(),
            voter.introducer.full_name if voter.introducer else '',
            voter.status or ''
        ])
    
    return response


@login_required
def export_candidates_csv(request):
    """تصدير المرشحين إلى CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="candidates.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'الكود', 'الاسم', 'الهاتف', 'البريد الإلكتروني',
        'المرتكزات', 'المعرّفين', 'الناخبين'
    ])
    
    for candidate in Candidate.objects.all():
        writer.writerow([
            candidate.candidate_code,
            candidate.full_name,
            candidate.phone,
            candidate.email or '',
            candidate.get_anchors_count(),
            candidate.get_introducers_count(),
            candidate.get_voters_count()
        ])
    
    return response


# ==================== Excel Export ====================

@login_required
def export_voters_excel(request):
    """تصدير الناخبين إلى Excel مع تنسيق احترافي"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "الناخبين"
    ws.right_to_left = True
    
    # تنسيق العنوان
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # العناوين
    headers = [
        'رقم الناخب', 'الاسم الكامل', 'تاريخ الميلاد', 'اسم الأم',
        'رقم الهاتف', 'المحافظة', 'مركز الاقتراع', 'رقم المركز',
        'رقم المحطة', 'التصنيف', 'المعرّف', 'الحالة'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # البيانات
    voters = Voter.objects.select_related('introducer').all()
    
    for row_num, voter in enumerate(voters, 2):
        ws.cell(row=row_num, column=1).value = voter.voter_number
        ws.cell(row=row_num, column=2).value = voter.full_name
        ws.cell(row=row_num, column=3).value = str(voter.date_of_birth) if voter.date_of_birth else ''
        ws.cell(row=row_num, column=4).value = voter.mother_name or ''
        ws.cell(row=row_num, column=5).value = voter.phone or ''
        ws.cell(row=row_num, column=6).value = voter.governorate
        ws.cell(row=row_num, column=7).value = voter.voting_center_name
        ws.cell(row=row_num, column=8).value = voter.voting_center_number
        ws.cell(row=row_num, column=9).value = voter.station_number or ''
        ws.cell(row=row_num, column=10).value = voter.get_classification_display()
        ws.cell(row=row_num, column=11).value = voter.introducer.full_name if voter.introducer else ''
        ws.cell(row=row_num, column=12).value = voter.status or ''
    
    # ضبط عرض الأعمدة
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="voters.xlsx"'
    wb.save(response)
    
    return response


@login_required
def export_daily_report_excel(request):
    """تقرير يومي شامل بصيغة Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    
    # صفحة الإحصائيات
    ws_stats = wb.active
    ws_stats.title = "الإحصائيات"
    ws_stats.right_to_left = True
    
    today = timezone.now().date()
    
    stats_data = [
        ['التقرير اليومي', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['', ''],
        ['إجمالي الناخبين', Voter.objects.count()],
        ['الناخبين المخصصين', Voter.objects.filter(introducer__isnull=False).count()],
        ['إجمالي المرشحين', Candidate.objects.count()],
        ['إجمالي المرتكزات', Anchor.objects.count()],
        ['إجمالي المعرّفين', Introducer.objects.count()],
        ['', ''],
        ['تصنيف الناخبين', ''],
        ['مؤيد', Voter.objects.filter(classification='supporter').count()],
        ['محايد', Voter.objects.filter(classification='neutral').count()],
        ['معارض', Voter.objects.filter(classification='opponent').count()],
        ['غير محدد', Voter.objects.filter(classification='unknown').count()],
        ['', ''],
        ['نشاط اليوم', ''],
        ['اتصالات اليوم', CommunicationLog.objects.filter(created_at__date=today).count()],
        ['مهام معلقة', CampaignTask.objects.filter(status='pending').count()],
        ['مهام مكتملة', CampaignTask.objects.filter(status='completed').count()],
    ]
    
    for row_num, (label, value) in enumerate(stats_data, 1):
        cell_label = ws_stats.cell(row=row_num, column=1)
        cell_value = ws_stats.cell(row=row_num, column=2)
        cell_label.value = label
        cell_value.value = value
        
        if row_num == 1:
            cell_label.font = Font(bold=True, size=14)
            cell_value.font = Font(bold=True, size=14)
        elif label and not value:
            cell_label.font = Font(bold=True)
    
    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 20
    
    # صفحة أعلى المراكز
    ws_areas = wb.create_sheet("أعلى المراكز")
    ws_areas.right_to_left = True
    ws_areas.append(['المركز الانتخابي', 'عدد الناخبين'])
    
    top_areas = Voter.objects.values('voting_center_name').annotate(
        count=Count('id')
    ).order_by('-count')[:20]
    
    for area in top_areas:
        ws_areas.append([area['voting_center_name'], area['count']])
    
    ws_areas.column_dimensions['A'].width = 50
    ws_areas.column_dimensions['B'].width = 15
    
    # الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="daily_report_{today}.xlsx"'
    wb.save(response)
    
    return response


# ==================== Dashboard للتقارير ====================

@login_required
def reports_dashboard(request):
    """لوحة تحكم التقارير"""
    context = {
        'excel_available': EXCEL_AVAILABLE,
        'pdf_available': PDF_AVAILABLE,
        'total_voters': Voter.objects.count(),
        'total_candidates': Candidate.objects.count(),
        'total_communications': CommunicationLog.objects.count(),
    }
    return render(request, 'elections/reports_dashboard.html', context)


# ==================== HTML Report ====================

@login_required
def daily_report_html(request):
    """تقرير يومي بصيغة HTML قابل للطباعة"""
    today = timezone.now()
    yesterday = today - timedelta(days=1)
    
    context = {
        'report_date': today,
        'total_voters': Voter.objects.count(),
        'assigned_voters': Voter.objects.filter(introducer__isnull=False).count(),
        'total_candidates': Candidate.objects.count(),
        'total_anchors': Anchor.objects.count(),
        'total_introducers': Introducer.objects.count(),
        
        # تصنيف
        'supporter_count': Voter.objects.filter(classification='supporter').count(),
        'neutral_count': Voter.objects.filter(classification='neutral').count(),
        'opponent_count': Voter.objects.filter(classification='opponent').count(),
        'unknown_count': Voter.objects.filter(classification='unknown').count(),
        
        # نشاط يومي
        'today_communications': CommunicationLog.objects.filter(
            created_at__date=today.date()
        ).count(),
        'yesterday_communications': CommunicationLog.objects.filter(
            created_at__date=yesterday.date()
        ).count(),
        
        # مهام
        'pending_tasks': CampaignTask.objects.filter(status='pending').count(),
        'in_progress_tasks': CampaignTask.objects.filter(status='in_progress').count(),
        'completed_tasks': CampaignTask.objects.filter(status='completed').count(),
        
        # أعلى المراكز
        'top_areas': Voter.objects.values('voting_center_name').annotate(
            count=Count('id')
        ).order_by('-count')[:10],
    }
    
    return render(request, 'elections/daily_report.html', context)


# ==================== تقارير المرشحين (Party Candidates) ====================

@login_required
def export_party_candidates_excel(request):
    """تصدير مرشحي الأحزاب إلى Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "مرشحو الأحزاب"
    ws.right_to_left = True
    
    # تنسيق العنوان
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # العناوين
    headers = [
        'اسم الحزب', 'الرقم التسلسلي للحزب', 'الرقم التسلسلي للمرشح',
        'الاسم الكامل', 'رقم الناخب', 'رقم الهاتف',
        'إجمالي الأصوات', 'تاريخ الإضافة'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # البيانات
    candidates = PartyCandidate.objects.select_related('party', 'voter').all()
    
    for row_num, candidate in enumerate(candidates, 2):
        ws.cell(row=row_num, column=1).value = candidate.party.name
        ws.cell(row=row_num, column=2).value = candidate.party.serial_number
        ws.cell(row=row_num, column=3).value = candidate.serial_number
        ws.cell(row=row_num, column=4).value = candidate.full_name
        ws.cell(row=row_num, column=5).value = candidate.voter_number or ''
        ws.cell(row=row_num, column=6).value = candidate.phone or ''
        ws.cell(row=row_num, column=7).value = candidate.get_total_votes()
        ws.cell(row=row_num, column=8).value = candidate.created_at.strftime('%Y-%m-%d')
    
    # ضبط عرض الأعمدة
    column_widths = [40, 20, 20, 35, 20, 20, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="party_candidates.xlsx"'
    wb.save(response)
    
    return response


@login_required
def export_party_candidates_pdf(request):
    """تصدير مرشحي الأحزاب إلى PDF"""
    if not PDF_AVAILABLE:
        return HttpResponse('مكتبة reportlab غير متوفرة', status=500)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="party_candidates.pdf"'
    
    # إنشاء المستند
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    
    # العنوان
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
    )
    
    title = Paragraph("تقرير مرشحي الأحزاب السياسية", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # جدول البيانات
    candidates = PartyCandidate.objects.select_related('party').all()[:100]  # الحد الأقصى 100 لتجنب PDF كبير جدًا
    
    table_data = [
        ['الحزب', 'رقم الحزب', 'رقم المرشح', 'الاسم', 'رقم الناخب', 'الأصوات']
    ]
    
    for candidate in candidates:
        table_data.append([
            candidate.party.name,
            str(candidate.party.serial_number),
            str(candidate.serial_number),
            candidate.full_name,
            candidate.voter_number or '-',
            str(candidate.get_total_votes())
        ])
    
    # إنشاء الجدول
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # بناء PDF
    doc.build(elements)
    
    return response

