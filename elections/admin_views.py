# -*- coding: utf-8 -*-
"""
Views خاصة بالأدمن لمراقبة مدراء المراكز
"""

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q, F, Sum
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from elections.models import (
    CenterDirector, AttendanceRecord, DirectorLoginLog,
    PoliticalEntityAgent, CandidateMonitor
)


def is_admin(user):
    """التحقق من أن المستخدم admin"""
    return user.is_authenticated and (user.is_superuser or user.is_staff)


@login_required
@user_passes_test(is_admin)
def admin_directors_monitor(request):
    """لوحة مراقبة مدراء المراكز (للأدمن فقط)"""
    
    # جميع المدراء
    directors = CenterDirector.objects.select_related('user', 'added_by').all()
    
    # إضافة إحصائيات لكل مدير
    for director in directors:
        # عدد مرات تسجيل الدخول
        director.login_count = DirectorLoginLog.objects.filter(
            director=director
        ).count()
        
        # آخر تسجيل دخول
        last_login = DirectorLoginLog.objects.filter(
            director=director
        ).order_by('-login_time').first()
        director.last_login = last_login
        
        # هل مسجل دخول حالياً؟
        director.is_logged_in = (
            last_login and 
            last_login.logout_time is None
        ) if last_login else False
        
        # عدد سجلات الحضور التي سجلها
        director.attendance_records_count = AttendanceRecord.objects.filter(
            recorded_by=director
        ).count()
        
        # عدد الحضور اليوم
        today = timezone.now().date()
        director.today_records = AttendanceRecord.objects.filter(
            recorded_by=director,
            created_at__date=today
        ).count()
        
        # عدد الوكلاء في مركزه
        director.agents_count = PoliticalEntityAgent.objects.filter(
            assigned_center_number=director.assigned_center_number
        ).count()
        
        # عدد المراقبين في مركزه
        director.monitors_count = CandidateMonitor.objects.filter(
            voting_center_number=director.assigned_center_number
        ).count()
    
    # إحصائيات عامة
    total_directors = directors.count()
    active_directors = sum(1 for d in directors if d.status == 'active')
    logged_in_now = sum(1 for d in directors if d.is_logged_in)
    total_login_logs = DirectorLoginLog.objects.count()
    total_attendance_records = AttendanceRecord.objects.count()
    
    # سجلات تسجيل الدخول الأخيرة
    recent_logins = DirectorLoginLog.objects.select_related(
        'director', 'user'
    ).order_by('-login_time')[:20]
    
    context = {
        'directors': directors,
        'total_directors': total_directors,
        'active_directors': active_directors,
        'logged_in_now': logged_in_now,
        'total_login_logs': total_login_logs,
        'total_attendance_records': total_attendance_records,
        'recent_logins': recent_logins,
        'page_title': 'مراقبة مدراء المراكز',
    }
    
    return render(request, 'elections/admin/directors_monitor.html', context)


@login_required
@user_passes_test(is_admin)
def attendance_reports(request):
    """تقارير الحضور والانصراف (للأدمن فقط)"""
    
    # تصفية حسب التاريخ
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    director_id = request.GET.get('director')
    record_type = request.GET.get('type')  # agent or monitor
    
    # جميع السجلات
    records = AttendanceRecord.objects.select_related(
        'recorded_by', 'agent', 'monitor'
    ).all()
    
    # تطبيق التصفيات
    if date_from:
        records = records.filter(created_at__date__gte=date_from)
    if date_to:
        records = records.filter(created_at__date__lte=date_to)
    if director_id:
        records = records.filter(recorded_by_id=director_id)
    if record_type:
        records = records.filter(record_type=record_type)
    
    records = records.order_by('-created_at')
    
    # إحصائيات
    total_records = records.count()
    checked_in = records.filter(status='checked_in').count()
    checked_out = records.filter(status='checked_out').count()
    
    # إحصائيات حسب المدير
    records_by_director = AttendanceRecord.objects.values(
        'recorded_by__full_name',
        'recorded_by__assigned_center_number'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # إحصائيات حسب النوع
    records_by_type = AttendanceRecord.objects.values(
        'record_type'
    ).annotate(
        count=Count('id')
    )
    
    # جميع المدراء (للتصفية)
    all_directors = CenterDirector.objects.all()
    
    context = {
        'records': records[:100],  # أول 100 سجل
        'total_records': total_records,
        'checked_in': checked_in,
        'checked_out': checked_out,
        'records_by_director': records_by_director,
        'records_by_type': records_by_type,
        'all_directors': all_directors,
        'filters': {
            'date_from': date_from,
            'date_to': date_to,
            'director': director_id,
            'type': record_type,
        },
        'page_title': 'تقارير الحضور والانصراف',
    }
    
    return render(request, 'elections/admin/attendance_reports.html', context)


@login_required
@user_passes_test(is_admin)
def export_attendance_excel(request):
    """تصدير تقرير الحضور إلى Excel"""
    
    # الحصول على التصفيات
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    director_id = request.GET.get('director')
    record_type = request.GET.get('type')
    
    # جميع السجلات
    records = AttendanceRecord.objects.select_related(
        'recorded_by', 'agent', 'monitor'
    ).all()
    
    # تطبيق التصفيات
    if date_from:
        records = records.filter(created_at__date__gte=date_from)
    if date_to:
        records = records.filter(created_at__date__lte=date_to)
    if director_id:
        records = records.filter(recorded_by_id=director_id)
    if record_type:
        records = records.filter(record_type=record_type)
    
    records = records.order_by('-created_at')
    
    # إنشاء Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير الحضور"
    
    # تنسيق الرأس
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # الرأس
    headers = [
        '#',
        'الاسم',
        'النوع',
        'الحالة',
        'وقت الحضور',
        'وقت الانصراف',
        'المدة',
        'المركز',
        'مدير المركز',
        'تاريخ التسجيل'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # البيانات
    for row_idx, record in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1).value = row_idx - 1
        ws.cell(row=row_idx, column=2).value = record.get_person_name()
        ws.cell(row=row_idx, column=3).value = record.get_record_type_display()
        ws.cell(row=row_idx, column=4).value = record.get_status_display()
        ws.cell(row=row_idx, column=5).value = record.check_in_time.strftime('%Y-%m-%d %H:%M') if record.check_in_time else '-'
        ws.cell(row=row_idx, column=6).value = record.check_out_time.strftime('%Y-%m-%d %H:%M') if record.check_out_time else 'لم ينصرف'
        ws.cell(row=row_idx, column=7).value = record.get_duration()
        ws.cell(row=row_idx, column=8).value = record.recorded_by.assigned_center_number
        ws.cell(row=row_idx, column=9).value = record.recorded_by.full_name
        ws.cell(row=row_idx, column=10).value = record.created_at.strftime('%Y-%m-%d %H:%M')
    
    # ضبط عرض الأعمدة
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # إنشاء Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def director_activity_detail(request, director_id):
    """تفاصيل نشاط مدير معين"""
    
    director = get_object_or_404(CenterDirector, id=director_id)
    
    # سجلات تسجيل الدخول
    login_logs = DirectorLoginLog.objects.filter(
        director=director
    ).order_by('-login_time')[:50]
    
    # سجلات الحضور
    attendance_records = AttendanceRecord.objects.filter(
        recorded_by=director
    ).select_related('agent', 'monitor').order_by('-created_at')[:100]
    
    # إحصائيات
    total_logins = DirectorLoginLog.objects.filter(director=director).count()
    total_attendance = AttendanceRecord.objects.filter(recorded_by=director).count()
    
    # آخر 7 أيام
    last_7_days = timezone.now().date() - timedelta(days=7)
    recent_attendance = AttendanceRecord.objects.filter(
        recorded_by=director,
        created_at__date__gte=last_7_days
    ).count()
    
    context = {
        'director': director,
        'login_logs': login_logs,
        'attendance_records': attendance_records,
        'total_logins': total_logins,
        'total_attendance': total_attendance,
        'recent_attendance': recent_attendance,
        'page_title': f'نشاط المدير: {director.full_name}',
    }
    
    return render(request, 'elections/admin/director_activity_detail.html', context)
